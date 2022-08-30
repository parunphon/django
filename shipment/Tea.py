#'======== coding: utf-8 ==================================================='
import sys
import os
import ntpath
import glob
import datetime
import time
import pandas as pd
import numpy as np
import configparser
import shutil
# conda update ipykernel
# conda clean --all
#conda update --all #need to set Control panel>Region>Administrative> set "English (US)"
#conda install -c anaconda pyodbc
#conda install -c anaconda sqlalchemy

#python -m pip install --upgrade pip
#pip install sqlalchemy
#pip install pyodbc
# to fix ImportError: Missing required dependencies ['pandas']
# pip uninstall pandas
# pip uninstall numpy
#Then restart comp. before reinstall
# pip install pandas
# pip install numpy 
# to fix ImportError: Missing required dependencies ['numpy']
# pip uninstall pandas
# pip uninstall numpy
# pip install pandas
# pip install numpy

# import py_compile
# script=r'D:\Siam City Cement Public Company Limited\Logistics_Reports - Documents\script\config.py'
# py_compile.compile(script)

# from flask import Flask
# from config import Config
# print('c',Config.BLOB_ACCOUNT)
# app = Flask(__name__)
# app.config.from_object(Config)
# print('a',app)
# print('b',app.config)
# print('e',app.config['BLOB_ACCOUNT'])
# print('f',app.config['TESTING'])

vmusername=['supplychainvm','win10','win102'] #,'ttuntiwi'
freight_users= vmusername+['win10']+ ['ttuntiwi','ajintana','jtraisor','naridbua','omeechak','asrangsr','prodwini','nauetavo','installwin7','installwin10' ,'tleelaha','pkornbon' ,'csajjanu','gjuntnap','t mankong','plonapal','ntrongji']
freight_columns=['IO Log region group','Existing Order Alloc Cost','Estimated Frieght Cost' 
                 ,'SuggestFreightCost','FinalFreightCost','billing_date','netweight_qty','swap_file','swap_cost'
                ,'basefreight_otm','accessorial_otm','labor_otm','cost_detail' ]
            #,'BASE_LOAD','BASE_EMPTY','OTHERS_LOAD','OTHERS_EMPTY','RENTAL','FUEL_SURCHARGE' ,'LABOR','RECYCLING','PORTFEE_PIPE','OTHERS']
freight_columns2=['BaseFreight_TripFreight','accessorial','labor','BahtPerTon','SOURCE_DATA']
#must not add 'Weight Out Date by SH',
################################ Gat USER NAME ########################################
computername=os.environ['COMPUTERNAME'].lower() #import platform ;platform.node() #import socket ;socket.gethostname()
username=ntpath.basename(os.environ['USERPROFILE']).lower()
#username='ttuntiwi' if username=='installwin10' else username
user=username[-8:]
isvmuser=True if username in vmusername else False
################################ Gat Data from Config File ##############################
#**********NEED TO RESTART when change value in config.ini**************
config_path=r'D:\Siam City Cement Public Company Limited\Logistics_Reports - Documents\script' if username in ['ttuntiwi'] else os.path.dirname(os.path.realpath(__file__)) 
config_pathfile=config_path+r'\config.ini'
contacts_pathfile=config_path+r'\contacts.xlsx' 
#if not use realpath when run in other environment ex.Task schedule, will give other folder.
config=configparser.ConfigParser()
config.read(config_pathfile)
print("Config file : ",config_pathfile,"\nContacts file : ",contacts_pathfile)
################################ Set Default value ########################################
programpath=os.path.abspath(os.path.join(os.path.dirname(__file__), '',))
#programpath=os.environ['USERPROFILE']+ r'\downloads\\'
print("Libraly path : ",programpath,"\nLibraly file : ",os.path.basename(__file__))

otm_user=config.get('OTM-config','otm_importdata_username')
das_user=config.get('das','username')
#otm_password=password_encode(config_pathfile,'OTM-config','otm_importdata_password') line 107
user_db=config.get('database','username')
user_onprimise=config.get('database','username_onprimise')
#--------------------------------------------------------------------
forcedb='sqldb-datawarehouse'
#--------------------------------------------------------------------
try: 
    try: defaultdb=config.get('database',computername)
    except Exception as e: defaultdb=config.get('database',username)
except Exception as e: defaultdb=config.get('database','default')
#--------------------------------------------------------------------
try: 
    try: sourcedb=config.get('database_source',computername)
    except Exception as e: sourcedb=config.get('database_source',username)
except Exception as e: sourcedb=config.get('database_source','default')
defaultdb='sqldb-datawarehouse' if sourcedb=='SCCC Data Warehouse' else defaultdb #for upload_missingdo()
print('Copmuter User Name:{} ,Current DB:{} ,Source DB:{}'.format(username,defaultdb,sourcedb))
###########################################################################################
def password_encode(configFilePath,config_group_name,config_password_name):
    import configparser
    from cryptography.fernet import Fernet
    #Fernet.generate_key()
    key =   b'sccc11112222333344445555666677888899990000x='
    cipher_suite = Fernet(key)
    ##### Get Data From Config File #####
    config = configparser.ConfigParser()
    #configFilePath = r'D:\PyThon\config.ini'
    config.read(configFilePath)
    password = config.get(config_group_name,config_password_name)
    if len(password) >= 100:
        password = bytes(password, "utf-8")
        password2 = cipher_suite.decrypt(password)
        password2 = password2.decode("utf-8")
        print("Password has been Encrypted;",password)
        password = password2
    else:
        cipher_text = cipher_suite.encrypt(password.encode())
        #Get the USERINFO section and Update the password
        userinfo = config[config_group_name]
        userinfo[config_password_name] = cipher_text.decode("utf-8")
        with open(configFilePath, 'w') as conf:
            config.write(conf)
            print("Password was not Encrypted --> already encoded;",cipher_text)    
    return password
otm_password=password_encode(config_pathfile,'OTM-config','otm_importdata_password')
das_password=password_encode(config_pathfile,'das','password')
password_db=password_encode(config_pathfile,'database','password')
password_onprimise=password_encode(config_pathfile,'database','password_onprimise')
###########################################################################################
def filter_columns(df):
    if df is None or df.empty: return df
    col_set=set(freight_columns2)
    user=ntpath.basename(os.environ['USERPROFILE'])
    if user.lower() not in freight_users:
        print('your user name:',user,'cannot see freight cost******')
        col_set=set(freight_columns+freight_columns2)
    col_list=[x for x in df.columns if x not in col_set]
    return df.loc[:,col_list]

def set_config_pathfile(active_dir):
    global config_pathfile
    config_pathfile=active_dir+r'\config.ini'
    return config_pathfile

def set_dbname(computername,config_pathfile):
    global defaultdb
    error=''
    try:
        config=configparser.ConfigParser()
        config.read(config_pathfile)
        defaultdb=config.get('database',computername)
    except Exception as e: error='(cannot load from config::{})'.format(e)
    print('*****************Set DB :"{}" from {}  {}*****************'.format(defaultdb,config_pathfile,error))
    return defaultdb

def get_user_password(system,config_pathfile=config_pathfile):
    username=config.get(system,'username')
    password=config.get(system,'password')
    server=config.get(system,'server')
    #server="smtp-mail.outlook.com" if server=='outlook' else "smtp.gmail.com"
    return [username,password,server]

def get_blob_connectionstring(containername,storageaccount='scccsupplychainstorage',config_pathfile=config_pathfile):
    storageaccountkey='' ;connectionstring=''
    if not os.path.isfile(config_pathfile): print('No file;',config_pathfile)
    else:
        config=configparser.ConfigParser()
        config.read(config_pathfile)
        if containername=='otm-data': storageaccount='scccsupplychainstorage'
        try:
            storageaccountkey=config.get(storageaccount,'storageaccountkey')
            connectionstring=config.get(storageaccount,'connectionstring')
        except Exception as e: print(e)
    return connectionstring,storageaccountkey 

class historySeries(pd.Series):
    @property
    def _constructor(self):
        return historySeries
    @property
    def _constructor_expanddim(self):
        return history
    def custom_series_function(self):
        return 'OK'
class history(pd.DataFrame): #['subject','detail','FileOutput','error','InsertUser','SendMail','RunTime']  
    # normal properties  print(pd.DataFrame._internal_names), properties defined in _metadata are retained, ex self.SUBJECT
    _metadata = ["_crs", "SUBJECT",'DETAIL','FILEOUTPUT','ERROR','SENDMAIL','RUNTIME','INSERTUSER']
    
    SUBJECT="subject"
    DETAIL="detail"
    FILEOUTPUT="FileOutput"
    ERROR="error"
    INSERTUSER="InsertUser"
    SENDMAIL="SendMail"
    RUNTIME="RunTime" 
    STARTTIME=datetime.datetime.now().strftime(" (%m%d_%H%M")
    DBNAME='sqldb-datawarehouse'
    
    def __init__(self,*args,**kwargs):
        crs = kwargs.pop("crs", None)
        kwargs['columns']=[self.SUBJECT,self.DETAIL,self.FILEOUTPUT,self.ERROR,self.SENDMAIL,self.RUNTIME]
        super(history, self).__init__(*args,**kwargs)
        #self.STARTTIME=datetime.datetime.now().strftime(" (%m%d_%H%M ")
        self._crs = crs if crs is not None else None
    @property
    def crs(self):
        return self._crs
    @property
    def _constructor(self):
        return history 
    def custom_frame_function(self):
        return 'OK'
    _constructor_sliced=historySeries
    
    def add(self,data_list):
        self.loc[len(self)]=data_list+[datetime.datetime.now()]
        print('Log added:',data_list)
        
    def save(self,savedir=config_path):
        #self.loc[:,'RunTime']=pd.to_datetime(self['RunTime'].apply(str),format='%y-%m-%d %H:%M:%S') 
        self.loc[:,self.INSERTUSER]=user
        self.loc[:,self.SUBJECT]=self[self.SUBJECT]+self.STARTTIME+datetime.datetime.now().strftime("_%H%M)_")+user
        try:
            self.loc[:,self.RUNTIME]=pd.to_datetime(self[self.RUNTIME]).dt.strftime('%Y-%m-%d %H:%M:%S')
            df_list=import_data_into_sql('','','INSERT_History_Daily',self,True) #True --> to get df from INSERT_History_Daily
            if isinstance(df_list,pd.DataFrame) and not df_list.empty:
                filename='History_1_{}.csv'.format(user)
                #fail ->df_list.loc[:,self.RUNTIME]=df_list[self.RUNTIME].dt.strftime('%Y-%m-%d %H:%M:%S')
                df_list.to_csv(os.path.join(savedir,filename),index=False,header=True)  #,sep='\t'
                print('Log saved to '+filename)
            elif isinstance(df_list,list) and df_list:
                for idx,df in enumerate(df_list):
                    filename='History_{}_{}.csv'.format(user,idx+1)
                    #fail ->if self.RUNTIME in df: df.loc[:,self.RUNTIME]=pd.to_datetime(self[self.RUNTIME]).dt.strftime('%Y-%m-%d %H:%M:%S') #df[self.RUNTIME].dt.strftime('%Y-%m-%d %H:%M:%S')
                    df.to_csv(os.path.join(savedir,filename),index=False,header=True)  #,sep='\t'
                    print('Log saved to '+filename)     
        except Exception as e: print('Cannot save {}::{}'.format(os.path.join(savedir,'History.csv'),e))             
#     def merge(self, *args, **kwargs):
#         #print(**kwargs)
#         #print(*args)
#         result = pd.DataFrame.merge(self,*args, **kwargs)
#         print(result)
#         result.__class__ = pd.DataFrame       
# #         geo_col = self._geometry_column_name
# #         if isinstance(result, DataFrame) and geo_col in result:
# #             result.__class__ = GeoDataFrame
# #             result.crs = self.crs
# #             result._geometry_column_name = geo_col
# #             result._invalidate_sindex()
# #         elif isinstance(result, DataFrame) and geo_col not in result:
# #             result.__class__ = DataFrame         
#         return result

def check_sql_string(sql,values_list):
    if values_list is not None:
        for v in values_list: sql = sql.replace('?',repr(v),1) #https://www.geeksforgeeks.org/str-vs-repr-in-python/
    return sql

def get_var_list(var_list1,var_list2):
    activedir_list,var_list2,BatchName=get_path_and_var(var_list1[0],var_list2)
    if len(activedir_list)==len(var_list1) :
        print('var_list_1 in batch file is equal to var_list_1 of program')
        var_list1=activedir_list
    return var_list1,var_list2
        
def checkpath():
    print('os.path.basename(__file__):',os.path.basename(__file__))
    print('os.path.realpath(__file__):',os.path.realpath(__file__))
    print('os.path.dirname(__file__):',os.path.dirname(__file__))
    print('os.path.dirname(os.path.realpath(__file__)):' ,os.path.dirname(os.path.realpath(__file__)))
    print('os.path.abspath(__file__):',os.path.abspath(__file__))
    print("os.path.join(os.path.dirname(__file__),'test'):",os.path.join(os.path.dirname(__file__),'test'))
    print("os.path.abspath(os.path.join(os.path.dirname(__file__),'test')):" ,os.path.abspath(os.path.join(os.path.dirname(__file__),'test')))

def get_path_and_var(var_list1, var_list2=None):
#def get_path_and_var(active_dir,var_list2=None, var_list2_start=2): 
    #sql_var_start starts from 0 #add this before import Tea : sys.path.append(os.environ['USERPROFILE']+r'\OTM_Project')
    import os
    import sys
    import ntpath
    print('------------------------------------------------------')
    ###---------------------------------
    active_dir=var_list1[0]
    var_list2_start=len(var_list1)
    ###---------------------------------
    python_exe=os.path.split(sys.executable)
    #----------------------------------------------
    batch_path='-' ;batch_name ='-' ;is_batch_path=False
    try: 
        if os.path.isdir(ntpath.split(sys.argv[1])[0]):
            batch_path, batch_name = ntpath.split(sys.argv[1])
            is_batch_path=True
        else:
            batch_path='-' ;batch_name ='-' ;is_batch_path=False
    except Exception as e: print("set batch_path='-',batch_name ='-', is_batch_path=False")

#     sys_argv='-f'
#     try: sys_argv=sys.argv[1:][0]
#     except Exception as e: print('error read argument::',e)
#     is_batch_path= False if sys_argv=='-f' else True
#     batch_path, batch_name = ntpath.split(sys_argv) if is_batch_path else ('-', "-")    
    batch_path=os.path.abspath(batch_path)
    #----------------------------------------------
    # ntpath.basename("a/b/c")
    script_path = os.getcwd() #os.path.split(os.getcwd())[0]
#     scriptname= sys.argv[1:][2] if is_batch_path else '-'
#     user_profile=os.environ['USERPROFILE']
#     tea_lib_path=user_profile+r'\OTM_Project'
    tea_lib_path=batch_path if is_batch_path else script_path
    
    #import inspect ;frame = inspect.stack()[1] ;script_path_name=frame[0].f_code.co_filename if is_batch_path else script_path+r'\\xxx.py'
    print('Python:        ',python_exe[0]+r'\\'+python_exe[1])
    if is_batch_path : print('Get parameter from Batch file : {}{}'.format(batch_path,'\\script\\'+batch_name))
    try: print('Library file:',tea_lib_path+r'\\script\\'+os.path.basename(__file__))
    except Exception as e: print('Library file error::',e)       
    #sys.path.append(r'C:\\Users\\TTUNTIWI')
    #sys.path.append(r'C:\Users\TTUNTIWI\OTM_Project')
    if script_path not in sys.path or tea_lib_path not in sys.path:
        sys.path.append(script_path)
        sys.path.append(tea_lib_path)
    if is_batch_path :
        result=[txt for txt in sys.argv[1:]] #unicode(sys.argv[1], 'utf-8')" or "sys.argv[1].decode('utf-8')
        #print('ttttt',[u'%s' %(txt,) for txt in sys.argv[1:]])
        result[0]=batch_path
    else : result=[active_dir]
    print(('Batch ' if is_batch_path else '')+'params :',result)  
#     sys.path.append(os.environ['USERPROFILE']+r'\OTM_Project') #***need if Tea is not in same path as batch file (and import sys, os)
    if len(result)>1 :
        active_dir=result[0] #batch always get act dir
        activedir_list=result[:var_list2_start]
        var_list2=result[var_list2_start:]
    else : activedir_list=result
    if len(activedir_list)==len(var_list1):
        var_list1=activedir_list
        print('var_list_1,2 : from batch file \'coz var_list_1 from batch file = from program')
    elif not var_list2 and len(activedir_list)>1: 
        var_list1=activedir_list
        print('*** get var_list1 : from batch file \'coz var_list1<>var_list2 but var_list2 is null')
    print('var_list_1:',var_list1,'\nvar_list_2:',var_list2,'\n---------------------------')
    return var_list1,var_list2,batch_name

#==SQL DATA WAREHOUSE ===================================================================
DB_params = "Driver={SQL Server};"\
            "Server=10.254.1.181;"\
            "Database={SCCC Data Warehouse};"\
            "uid=sa;pwd=test"
            #"autocommit=True"
            #"Driver={SQL Server Native Client 11.0};"\
def get_db_params(db_name):
    if db_name in ['SCCC Data Warehouse']:
        credentials = { 'username'  : user_onprimise
                        ,'password'  : password_onprimise
                        ,'host'      : '10.254.1.181'
                        ,'database'  : db_name
                        ,'driver'    : 'SQL Server'}
    elif db_name in ['sqldb-datawarehouse','sqldb-gps_silo']:
        credentials = { 'username'  : user_db
                        ,'password'  : password_db
                        ,'host'      : 'sql-supplychain.database.windows.net'
                        ,'database'  : db_name
                        ,'driver'    : 'ODBC Driver 17 for SQL Server'} #'SQL Server Native Client 11.0','ODBC Driver 17 for SQL Server'
    elif db_name in ['sqldb-testdev']:
        credentials = { 'username'  : 'readonly'
                        ,'password'  : 'Warehouse8*'
                        ,'host'      : 'sql-supplychain.database.windows.net'
                        ,'database'  : db_name
                        ,'driver'    : 'ODBC Driver 17 for SQL Server'} #'SQL Server Native Client 11.0','ODBC Driver 17 for SQL Server'
    elif db_name in ['SCCCLogisticsDataWarehouse']:
        credentials = { 'username'  : user_db
                        ,'password'  : password_db
                        ,'host'      : 'sccclogistic.database.windows.net'
                        ,'database'  : db_name
                        ,'driver'    : 'ODBC Driver 17 for SQL Server'} #'SQL Server Native Client 11.0','ODBC Driver 17 for SQL Server'
    elif db_name in ['SCCCLogistic']:
        credentials = { 'username'  : user_db
                        ,'password'  : password_db
                        ,'host'      : 'sccclogistic.database.windows.net'
                        ,'database'  : db_name
                        ,'driver'    : 'ODBC Driver 17 for SQL Server'}  #'SQL Server Native Client 11.0','ODBC Driver 17 for SQL Server'
    elif db_name=='SCCC_Tracking':
        credentials = { 'username'  : 'sccc_tracking'
                        ,'password'  : 'sccc_tracking'
                        ,'host'      : '210.1.60.112,1433'
                        ,'database'  : db_name
                        ,'driver'    : 'SQL Server'}
    else: 
        print('No database:',db_name)
        return None
    print('Connect DB: {}, host: {}, driver: {}'.format(db_name,credentials['host'],credentials['driver']))
    return credentials
# https://www.connectionstrings.com/download-sql-server-native-client/
# https://docs.microsoft.com/en-us/sql/connect/odbc/download-odbc-driver-for-sql-server?view=sql-server-ver15
# https://docs.microsoft.com/th-th/office/troubleshoot/access/sql-server-native-client-drivers
# http://go.microsoft.com/fwlink/?LinkID=239648&clcid=0x409
# {SQL Server} - released with SQL Server 2000
# {SQL Native Client} - released with SQL Server 2005 (also known as version 9.0)
# {SQL Server Native Client 10.0} - released with SQL Server 2008
# {SQL Server Native Client 11.0} - released with SQL Server 2012
# {ODBC Driver 11 for SQL Server} - supports SQL Server 2005 through 2014
# {ODBC Driver 13 for SQL Server} - supports SQL Server 2005 through 2016
# {ODBC Driver 13.1 for SQL Server} - supports SQL Server 2008 through 2016
# {ODBC Driver 17 for SQL Server} - supports SQL Server 2008 through 2019

def connectDB_(db_name):
    import sqlalchemy
    credentials=get_db_params(db_name)
    connect_url='' ;engine=None
    if not credentials: print ("Error, did not connect DB; no credentials")
    else:
        try:
            connect_url = sqlalchemy.engine.url.URL(
                'mssql+pyodbc'
                ,username=credentials['username']
                ,password=credentials['password']
                ,host=credentials['host'] #port=credentials['port'],
                ,database=credentials['database']
                ,query=dict(driver=credentials['driver']))
            engine = sqlalchemy.create_engine(connect_url) #,echo=True #connection=engine.connect()
        except Exception as e: error=str(e)
#     engine = sqlalchemy.create_engine('mssql+pyodbc://sa:xxxxx@10.254.1.181/SCCC Data Warehouse?driver=SQL+Server+Native+Client+11.0')
#     print(engine.url)
    if engine is None or not engine: print('Error, did not connect DB by sqlalchemy.engine.url.URL(credentials)')
    return engine

def connectDB(db_name):
    import pyodbc
    credentials=get_db_params(db_name)
    conn=False ;error=''
    if not credentials: print ("Error, did not connect DB; no credentials")
    else:
        try: #Do not put a space after the Driver keyword in the connection string
            conn_str='DRIVER={'+credentials['driver']+'};'\
                      'SERVER='+credentials['host']+';DATABASE='+credentials['database']+';'\
                      'UID='+credentials['username']+';PWD='+ credentials['password'] +';' #'Trusted_Connection=yes;' #'autocommit=True')
            conn = pyodbc.connect(conn_str) 
        except Exception as e: error=str(e)
#     if __name__ == "__main__": conn = conn #connect db
    if not conn or error:
        error = f'{conn}: did not connect DB by pyodbc.connect(conn_str):: {error}'
        print(error)
    print('-------------', conn, '----------------')
    return conn, error
#--------------------------------------------------------------------------------------------
def store_var(store_name=''):
    x=None ;store_name=store_name.lower()
    if store_name[:16] in 'search_shipment_report': #store_name[:16]='search_shipment_'
        x={'WO_DateFrom':'','WO_DateTo':''
             ,'PreDOCreate_DateFrom':''   ,'PreDOCreate_DateTo':''
             ,'ShipmentCreate_DateFrom':'','ShipmentCreate_DateTo':''
             ,'RequestTo_DateFrom':''     ,'RequestTo_DateTo':''
             ,'LogConfirmTo_DateFrom':''  ,'LogConfirmTo_DateTo':''
             ,'Assign_DateFrom':''        ,'Assign_DateTo':''
             ,'ArriveAtCust_DateFrom':''  ,'ArriveAtCust_DateTo':''
             ,'LogConfirmPOD_DateFrom':'' ,'LogConfirmPOD_DateT':''
             ,'ShipmentID':'','PreDONo':'' ,'DONo':'' ,'PONo':''
             ,'Perspective':'' ,'DFM':'' ,'MultiStopMark':''
             ,'OrderType':''        ,'ShipmentStatus':''
             ,'ShippingCondition':'','ShippingType':''
             ,'LogRegionGroup':''   ,'LogRegionCode':'','IOLogRegionGroup':''
             ,'TZoneDescription':'' ,'ProvinceDescription':''
             ,'BHIndex':''          
             ,'SoldToCode':''       ,'SoldToName':''
             ,'ShipToCode':''       ,'ShipToName':''
             ,'LogMatGroup':''      
             ,'Plant':'' ,'ShippingPoint':'' 
             ,'Material':''         ,'MatDescription':''
             ,'TruckID':''          ,'Driver':''
             ,'TransporterCode':''  ,'TransporterName':''
             ,'ConditionGroup5':''  ,'AssignUser':''
             ,'Division':''         ,'SalesOrg':''
             ,'mark_code':''        ,'Delivery_note':''
             ,'ilm_DateFrom':''     ,'ilm_DateTo':''
             ,'gps_result':''       ,'ontime':'' 
             ,'Is_byShipment':''    
             ,'GroupBy_Sum':'' ,'GroupBy_Rows':'' ,'GroupBy_Columns':''}    
    elif store_name in ['search_master','search_history']: x={'table':'','doc_no':'','doc_name':'','DateFrom1':'','DateTo1':''}    
    elif store_name in 'validate_shipment_cost': x={'LogConfirmPOD_DateFrom':'' ,'LogConfirmPOD_DateTo':''}    
    elif store_name in ['kfa_select_data','kyw_select_data']: x={'IsHistory':'','DateTo':'' ,'plant':'','rank':''}    
    elif 'upsert' in store_name: x={'source':None}    
    else: print('No variable to execute')
    return x

def store_query(exec='',sql_var=None,is_print_query=True): #if no paremeter, set to None
    exec=exec.lower()
    jsonquery_list=['calculate_missingdo_temp','insert_freight_master_temp','insert_shipment_ontime_temp' ,'insert_shipmenttracking_temp' ,'update_do_temp','update_location_route_temp','insert_history_daily_temp','insert_das_weight_temp']
    #exec='{call dbo.'+store_exec+'(@TruckID=?,@DateFrom=?)}'  #ODBC format cannot get output param 
    if exec in jsonquery_list: exec=f'EXEC dbo.{exec} @db=? ,@json=?'
    elif exec in ['search_master','search_history']:
        exec='EXEC dbo.'+exec+' \n' \
        "@table=?        ,@doc_no=?       ,@doc_name=? \n" \
        ",@DateFrom1=?   ,@DateTo1=? \n"
    elif exec=='search_tracklog': 
        exec="DECLARE @out nvarchar(max); \n"\
        'EXEC dbo.'+exec+' \n'\
        "@TruckID=? ,@DateFrom=? ,@DaysAdd=? \n"\
        ",@param_out=@out OUTPUT; \n"\
        "SELECT @out AS the_output;"
    elif exec=='check_weight_diff':
        exec='DECLARE @out nvarchar(max); \n'\
        'EXEC dbo.'+exec+' @param_out=@out OUTPUT; \n'\
        'SELECT @out AS the_output;'
    elif exec in ['search_shipment_report_ontime']:
        exec='EXEC dbo.'+exec
    elif exec in ['search_shipment_report','search_shipment_report_ilm','search_shipment_report_freight']:
        exec="DECLARE @out nvarchar(max); \n"\
        'EXEC dbo.'+exec+' \n'\
        "@WO_DateFrom=?             ,@WO_DateTo=? \n"\
        ",@PreDOCreate_DateFrom=?   ,@PreDOCreate_DateTo=? \n"\
        ",@ShipmentCreate_DateFrom=?,@ShipmentCreate_DateTo=? \n"\
        ",@RequestTo_DateFrom=?     ,@RequestTo_DateTo=? \n"\
        ",@LogConfirmTo_DateFrom=?  ,@LogConfirmTo_DateTo=? \n"\
        ",@Assign_DateFrom=?        ,@Assign_DateTo=? \n"\
        ",@ArriveAtCust_DateFrom=?  ,@ArriveAtCust_DateTo=? \n"\
        ",@LogConfirmPOD_DateFrom=? ,@LogConfirmPOD_DateTo=? \n"\
        ",@ShipmentID=? ,@PreDONo=? ,@DONo=? ,@PONo=? \n"\
        ",@Perspective=? ,@DFM=? ,@MultiStopMark=? \n"\
        ",@OrderType=?              ,@ShipmentStatus=? \n"\
        ",@ShippingCondition=?      ,@ShippingType=? \n"\
        ",@LogRegionGroup=?         ,@LogRegionCode=?      ,@IOLogRegionGroup=? \n"\
        ",@TZoneDescription=?       ,@ProvinceDescription=? \n"\
        ",@BHIndex=? \n"\
        ",@SoldToCode=?             ,@SoldToName=?\n"\
        ",@ShipToCode=?             ,@ShipToName=? \n"\
        ",@LogMatGroup=? \n"\
        ",@Plant=?                  ,@ShippingPoint=? \n"\
        ",@Material=?               ,@MatDescription=? \n" \
        ",@TruckID=?                ,@Driver=? \n"\
        ",@TransporterCode=?        ,@TransporterName=? \n"\
        ",@ConditionGroup5=?        ,@AssignUser=? \n"\
        ",@Division=?               ,@SalesOrg=? \n"\
        ",@mark_code=?              ,@Delivery_note=? \n"\
        ",@ilm_DateFrom=?           ,@ilm_DateTo=? \n"\
        ",@gps_result=?             ,@ontime=? \n"\
        ",@Is_byShipment=? \n"\
        ",@GroupBy_Sum=? ,@GroupBy_Rows=? ,@GroupBy_Columns=? \n"\
        ",@param_out=@out OUTPUT; \n"\
        "SELECT @out AS the_output;"
    elif exec in ['kfa_select_data','kyw_select_data']:
        exec="EXEC dbo."+exec+" \n"\
        "@IsHistory=?   ,@DateTo=? \n"\
        ",@plant=?       ,@rank=?"
        # y='SCCCLogistic' if exec=='kfa_select_data'else 'SCCC_Tracking'
    elif exec == 'validate_shipment_cost':
        exec = '{call dbo.'+exec+'(@LogConfirmTo_DateFrom=?,@LogConfirmTo_DateTo=?)}'
    elif 'upsert' in exec: exec="EXEC dbo."+exec+" @source=?"
    else: print('No store procedure: '+exec)
    return exec
    
#     import pandas as pd
#     #---1. Set parameters for query
#     params=sql_var if sql_var is None or isinstance(sql_var,list) else list(sql_var.values()) #sql_var.items()
#     #print(params)
#     #---2. Print & execute query
#     conn=connectDB(y)#.connect()
#     cursor=conn.cursor()
#     if params is None : rows=cursor.execute(x).fetchall()
#     else : 
#         rows=cursor.execute(x,params).fetchall() #SQL Server format-->#while rows: print(rows)
#         query=check_sql_string(x,params)
        
#     if is_print_query: print("------Execute query----- \n",x)
#     #-------------------------------------
#     columns=[column[0] for column in cursor.description] #must before move cursor
#     df=pd.DataFrame.from_records(rows,columns=columns) #df=pd.read_sql(sql=query,con=conn,params=params)
    
#     if is_print_query and cursor.nextset() : print('------Query -----',cursor.fetchall()[0][0])
#     cursor.close()
#     conn.close()
#     return df
def sql_execute_query_(db_name,query,sql_var=None): #if no paremeter, set to None
    #***cannot execute query/storeprocedure without output
    import pandas as pd
    #---1. Set parameters for query
    params=sql_var if sql_var is None or isinstance(sql_var,list) else list(sql_var.values()) #sql_var.items()
    
    #---2. Print & execute query
    engine = connectDB_(db_name) 
    conn = engine.raw_connection()#engine.connect() #engine.raw_connection().connection
    cursor = conn.cursor()
    
    if params is None : rows=cursor.execute(query) #rows=conn.execute(query)
    else : 
        rows=cursor.execute(query,params) #conn.execute(query,params)
        query=check_sql_string(query,params)
    
    print("------Execute query----- \n",query)
    #-------------------------------------
    columns=[column[0] for column in cursor.description] #must before move cursor
    df=pd.DataFrame.from_records(rows,columns=columns) #df=pd.read_sql(sql=query,con=conn,params=params)
    #for x in rows: print("shipmentID:", x['shipmentID'])
    if cursor.nextset() : print('------Query -----',cursor.fetchall()[0][0])
    
    cursor.commit() #need
    cursor.close()
    conn.close()
    return df 

def sql_execute_query(db_name,query,sql_var=None,is_print_query=False): #if no paremeter, set to None
    import pandas as pd
    #---1. Set parameters for query
    if query=='' or db_name=='': return None #if sql_var={}, sql_var.values()=[]
    sql_var=None if isinstance(sql_var,dict) and not sql_var else sql_var
    params=list(sql_var.values()) if isinstance(sql_var,dict) else sql_var #sql_var.items()
    df=None
    #---execute sqldb-datawarehouse ------------------------------------------------
#     if 'delete_predo_shipment' in query.lower():
#         if db_name=='sqldb-datawarehouse':
#             df=sql_execute_query('SCCC Data Warehouse',query,sql_var,is_print_query)
#             #print('yyyyyyyy Complete {} in SCCC Data Warehouse'.format(query[:200])) 
#             return df
    #---2. Print & execute query
    conn, error = connectDB(db_name)  # .connect()
    if error: return df
    try:
        cursor=conn.cursor()
        if params is None : rows=cursor.execute(query).fetchall()
        else: rows=cursor.execute(query,params).fetchall() #SQL Server format-->#while rows: print(rows)
        #print('yyyyyyyy Complete {}...'.format(query[:200]))    
        #-------------------------------------
        columns=[column[0] for column in cursor.description] #must before move cursor
        df=pd.DataFrame.from_records(rows,columns=columns) #df=pd.read_sql(sql=query,con=conn,params=params)
        
        if is_print_query:
            print("------Execute query----- \n",check_sql_string(query,params))
            if cursor.nextset(): 
                try: print('------Query -----',cursor.fetchall()[0][0])
                except Exception as e: 
                    print('zzzzzzzz Error when print query',e)
                    return None

        cursor.commit() #need when update db ,ex.'delete_predo_shipment'
        cursor.close()
    except Exception as e: print('zzzzzzzz Error; cannot execute {} :: {}'.format(query,e))    
    finally: conn.close()
    df=filter_columns(df)
    return df

def save_json(path,filename,df):
    if isinstance(path,str): df.to_json(r'{}\{}.json'.format(path,filename),orient='records')
    else: #path=containerClient
        blob=path.get_blob_client(filename+'.json',snapshot=None)
        blob.upload_blob(df.to_json(orient='records'),overwrite=True)

def import_data_into_sql(active_dir,table_list,store_execute_list,filename_list=None ,is_return_df_from_store=False ,if_exists='append' ,**read_excel_kwargs):
    import datetime
    df=pd.DataFrame() ;df_list=[] ;filename='' ;time_str='_'+user+datetime.datetime.now().strftime('_%d%H%M%S')
    if isinstance(store_execute_list,str): store_execute_list=[] if not store_execute_list else store_execute_list.strip().split(',') 
    if isinstance(filename_list,str): filename_list=filename_list.strip().split(',') 
    if isinstance(filename_list,pd.DataFrame): filename_list=[filename_list]
    store_execute_list=[x+'_temp' if '_temp' not in x.lower() else x for x in store_execute_list]
    if not store_execute_list: return filename_list[0]
    if len(filename_list)!=len(store_execute_list):
        print('cannot load data:{} input files <> {} executing store precedures'.format(len(filename_list),len(store_execute_list)))
        return [df]*len(store_execute_list)
    dir_name=active_dir if isinstance(active_dir,str) else 'ContainerClient'
    conn,error=connectDB(defaultdb)
    if error: return df_list[0] if len(df_list)==1 else df_list
    try:
        cursor=conn.cursor() ;df_json=''
        for n,storename in enumerate(store_execute_list):
            filename=storename+time_str
            try:
                if isinstance(filename_list[n],str) and os.path.splitext(filename_list[n])[1].lower()=='.json':
                    if isinstance(active_dir,str):
                        f=open(os.path.join(active_dir,filename_list[n]), "r")
                        df_json=f.read() #with open(os.path.join(active_dir,filename_list[n])) as json_file: df_json = json.load(json_file)
                    else:
                        blob=active_dir.get_blob_client(filename_list[n],snapshot=None)
                        df_json=blob.download_blob().readall() #df=pd.read_json(df_json)
                else:
                    df=filename_list[n] if isinstance(filename_list[n],pd.DataFrame) else get_df_from_df_latest_file(active_dir,filename_list[n] ,**read_excel_kwargs)
                    if sourcedb!='otm': DateTime_ToString(df,'%Y-%m-%d %H:%M:%S')
                    df.loc[:,'INSERT_BY']=user ;df.loc[:,'UPDATE_BY']=user
                    df_json=df.to_json(orient='records') #orient=split,records,index,values,table,columns(default format)

                params=[sourcedb,df_json] ;params2=[sourcedb,df_json[:2500]]
                query='EXEC dbo.{} @db=? ,@json=?'.format(storename) 
                script='' if 'history_daily' in storename.lower() else query
                for v in params2: script=script.replace('?',repr(v),1) #https://www.geeksforgeeks.org/str-vs-repr-in-python/
                print('{} exucute store procedure {}\n{}...=================='.format(n+1,storename,script))
                complete_word='' if 'history_daily' in storename.lower() else 'zzzzzzzz'
                if not is_return_df_from_store:
                    rows=cursor.execute(query,params) #cursor.execute(query,params).fetchall()
                    print('{} Complete {}'.format(complete_word,storename))
                    df_list.append(df)
                else: # get First
                    rows=cursor.execute(query,params).fetchall()
                    columns=[column[0] for column in cursor.description] #must before move cursor
                    df=pd.DataFrame.from_records(rows,columns=columns)
                    df_list.append(df)
                    print('{} Complete {} and got the 1st returned table'.format(complete_word,storename))
                    try:# check for more results
                        i=2
                        while (cursor.nextset()):
                            rows=cursor.fetchall() #*** need to re-fetchall 'coz 2nd table may have different columns
                            columns=[column[0] for column in cursor.description]
                            df=pd.DataFrame.from_records(rows,columns=columns)
                            df_list.append(df)
                            print('{} Complete {} and got the {}nd returned table'.format(complete_word,storename,i))
                            i+=1
                    except Exception as e: print('{} has no more {}rd returned table:: {}'.format(storename,i,e))
                #============================================================
                cursor.commit() #need
                if len(store_execute_list)>1: time.sleep(3)
            except Exception as e:
                print('zzzzzzzz Error execute {} and save to {}\{}.json::{}'.format(storename,dir_name,filename,e))
                if isinstance(df,pd.DataFrame) and not df.empty: save_json(active_dir,filename,df)
                df_list.append(pd.DataFrame()) #empty df #sys.exit(1)
        
        # cursor.close()  #remove on 21/6/2021 # rows.close()
    except Exception as e:
        print("zzzzzzzz Error execute {}::{}".format(store_execute_list,e))
        df_list=df_list+[pd.DataFrame()]*(len(store_execute_list)-len(df_list)) #empty df #sys.exit(1)
    finally: 
        conn.close()
        print('=========================Done with {} records ================================='.format(len(df)))
        return df_list[0] if len(df_list)==1 else df_list

def import_otm_into_sql_table(database,active_dir,table_list,store_execute_list,input_list=None ,is_return_df_from_store=False ,if_exists='append' ,**read_excel_kwargs):
    df=pd.DataFrame() ;df_list=[] ;filename='' ;time_str='_'+user+datetime.datetime.now().strftime('_%d%H%M%S')
    if isinstance(store_execute_list,str): store_execute_list=store_execute_list.strip().split(',') 
    if isinstance(input_list,str): input_list=input_list.strip().split(',') 
    if isinstance(input_list,pd.DataFrame): input_list=[input_list]
    isSingleDf=len(store_execute_list)==1
    store_execute_list=[x+'_temp' if '_temp' not in x.lower() else x for x in store_execute_list]
    if len(input_list)!=len(store_execute_list):
        print('cannot load data:{} input files <> {} executing store precedures'.format(len(input_list),len(store_execute_list)))
        return df if isSingleDf else [df]*len(store_execute_list)
    dir_name=active_dir if isinstance(active_dir,str) else 'ContainerClient'
    conn,error=connectDB(database)
    if error: return df if isSingleDf else df_list
    try:
        cursor=conn.cursor() ;df_json=''
        for n,storename in enumerate(store_execute_list):
            filename=storename+time_str
            try:
                if isinstance(input_list[n],str) and os.path.splitext(input_list[n])[1].lower()=='.json':
                    if isinstance(active_dir,str):
                        f=open(os.path.join(active_dir,input_list[n]), "r")
                        df_json=f.read() #with open(os.path.join(active_dir,input_list[n])) as json_file: df_json = json.load(json_file)
                    else:
                        blob=active_dir.get_blob_client(input_list[n],snapshot=None)
                        df_json=blob.download_blob().readall() #df=pd.read_json(df_json)
                else:
                    df=get_df_from_df_latest_file(active_dir,input_list[n] ,**read_excel_kwargs)
                    if sourcedb!='otm': DateTime_ToString(df,'%Y-%m-%d %H:%M:%S')
                    df.loc[:,'INSERT_BY']=user ;df.loc[:,'UPDATE_BY']=user
                    df_json = df.to_json(orient='records') #orient=split,records,index,values,table,columns(default format)

                params=[sourcedb,df_json] ;params2=[sourcedb,df_json[:2000]]
                query=store_query(storename) ;script='' if 'history_daily' in storename.lower() else query
                for v in params2: script=script.replace('?', repr(v), 1) #https://www.geeksforgeeks.org/str-vs-repr-in-python/
                print('{} exucute store procedure {}================== \n{}...'.format(n+1,storename,script))

                rows=cursor.execute(query,params)#.fetchall()
                print('{} Complete {}'.format('' if 'history_daily' in storename.lower() else 'zzzzzzzz',storename))
                if is_return_df_from_store and n==len(store_execute_list)-1: 
                    columns=[column[0] for column in cursor.description] #must before move cursor
                    df=pd.DataFrame.from_records(rows,columns=columns)
                cursor.commit() #need
                if len(store_execute_list)>1: time.sleep(3)
            except Exception as e:
                print('zzzzzzzz Error execute {} and save to {}\{}.json::{}'.format(storename,dir_name,filename,e))
                if isinstance(df,pd.DataFrame) and not df.empty: save_json(active_dir,filename,df)
                df=pd.DataFrame() #empty df #sys.exit(1)
            finally: df_list.append(df)
        cursor.close() #rows.close()
    except Exception as e:
        print("zzzzzzzz Error execute {}::{}".format(store_execute_list,e))
        df_list=df_list+[pd.DataFrame()]*(len(store_execute_list)-len(df_list)) #empty df #sys.exit(1)
    finally: 
        conn.close()
        print('=========================Done with {} records ================================='.format(len(df)))
        return df if isSingleDf else df_list

def import_otm_into_sql_table_(database,active_dir,temp_table_list,store_execute_list ,input_list=None,is_return_df_from_store=False ,if_exists='append' ,**read_excel_kwargs):
    import pandas as pd
    time_str=''
    if isinstance(temp_table_list,str): temp_table_list=temp_table_list.strip().split(',') 
    if isinstance(store_execute_list,str): store_execute_list=store_execute_list.strip().split(',') 
    if isinstance(input_list,str): input_list=input_list.strip().split(',') 
    if isinstance(input_list,pd.DataFrame): input_list=[input_list]
    if len(input_list)!=len(temp_table_list):
        print('cannot load data:{} input files list <> {} temp tables list'.format(len(input_list),len(temp_table_list)))
        return
    try:
        engine=connectDB_(database) #connection=engine.connect()
        cursor = engine.raw_connection().cursor()
        m=0
        for x in temp_table_list:
            if x=='': continue
            if 'temp' in x.lower():
                print('#1.'+str(m+1)+' Delete temp table ('+x+')===========================')
                query='DELETE FROM '+x+'; \n'
                query+='SET dateformat dmy' #prepare to insert
                rows=engine.execute(query) #sql_execute_query_(query_delete,None)
                
            print('#2. Insert Converted file into SQL table; '+x+'======')
            df=get_df_from_df_latest_file(active_dir,input_list[m],**read_excel_kwargs)
            df.to_sql(x,con=engine,chunksize=20,method='multi',index=False,if_exists=if_exists) 
            #'replace': Drop the table before inserting new values.
            #‘multi’: Pass multiple values in a single INSERT clause, it significantly reduces the task execution time.
            #https://stackoverflow.com/questions/25661754/get-data-from-pandas-into-a-sql-server-with-pyodbc
            #input_list can be list of 2 members as per get_df_from_df_latest_file()
            #17 oct ; try to map df with temp_table_list to delete and insert  but still cannot
            filename=x+datetime.datetime.now().strftime('_%d%H%M%S')
            df.to_json(r'{}\{}.json'.format(active_dir,filename),orient='records')
            m+=1         
        n=0
        for y in store_execute_list:
            if y=='': continue
            print('#4.'+str(n+1)+' Exucute temp into SQL table by script; '+y+'==================')
            ##query="dbo.INSERT_ShipmentTracking" rows=engine.execute(query) #cannot
            ##sql_execute_query_(query,None)
            query='{CALL dbo.'+y+'}'
#             cursor.execute(query)
            rows=cursor.execute(query)#.fetchall()
            if is_return_df_from_store and n==len(store_execute_list)-1: 
                columns=[column[0] for column in cursor.description] #must before move cursor
                df=pd.DataFrame.from_records(rows,columns=columns)
            cursor.commit() #need
            n+=1
       #rows.close()
        cursor.close()
        print('==================Done===================================')
    except KeyError:
        print("No data input ")
        return False
        sys.exit(1)       
    return df

def load_data(path, filelistname, sheetname=None, skiprows=0,csv_sep=',',is_setdf=False,**read_excel_kwargs): 
    #read_exce :,converters={'names':str,'ages':str} ||| read_csv :,dtype={csv_id:str} or dtype=str If converters are specified, they will be applied INSTEAD of dtype conversion.
    #if sheetname=None or many sheets in Excel file--> result is data dict/orderDict
    import collections
    df=None
    if isinstance(path,str):    
        ###cwd = os.getcwd() # Retrieve current working directory (`cwd`) ;os.listdir()  # List all files and directories in current directory
        #os.chdir(path) ;filelist = glob.glob('*{}*'.format(filelistname)) if type(filelistname)==str else filelistname
        #filelist=[x for x in filelist if x[0].isalnum()]
        filelist=[x for x in all_files_under(path,filelistname,True,'in','filename')] if isinstance(filelistname,str) else filelistname
        filelist=[file for file in filelist if any(word in file for word in ['.xls','.csv'])]
        if len(filelist)<1: 
            print('No file {}'.format(os.path.join(path,filelistname)))
            return None
    #     #SORT FILES: leave only regular files, insert creation date
    #     #NOTE: on Windows `ST_CTIME` is a creation date. But on Unix it could be something else. Beter use `ST_MTIME` to sort by a modification date
    #     from stat import S_ISREG, ST_CTIME, ST_MTIME, ST_MODE
    #     filelist=((os.stat(path), path) for path in filelist) #ctime=datetime.date.fromtimestamp(os.path.getctime(active_dir+r'\\'+filename))
    #     filelist=((stat[ST_CTIME], path) for stat, path in filelist if S_ISREG(stat[ST_MODE]))
    #     filelist=[path for cdate, path in sorted(filelist)] #os.path.basename(path)

        if sheetname=='': sheetname=None
        pre_sheetname=sheetname
        for x in ['']:
            if x in read_excel_kwargs: read_excel_kwargs.pop(x)
        for idx, filename in enumerate(filelist):
            file_extension = os.path.splitext(filename)[1].lower()
            is_csv=file_extension in ('.csv','.xls') ;is_otm=file_extension=='.xls'
            try:
                if is_csv:       
                    if is_otm:
                        csv_sep='\t'
                        dfsub=pd.read_csv(path+os.path.sep+filename,skiprows=skiprows,sep=csv_sep,low_memory=False,encoding="UTF-16 LE",**read_excel_kwargs)
                    else: dfsub=pd.read_csv(path+os.path.sep+filename,skiprows=skiprows,sep=csv_sep,low_memory=False,**read_excel_kwargs)
                    #dfsub=pd.read_table(filename, encoding='utf-16')
                else:
                    sheet_list=get_all_sheets(path,filename,'')
                    if len(sheet_list)==1: sheetname=sheet_list[0] # and len(filelist)==1
                    # dfsub=pd.read_excel(path+os.path.sep+filename,sheet_name=sheetname,skiprows=skiprows,na_values=['','NA','nan','null'],**read_excel_kwargs)
                    if sheetname and sheetname not in sheet_list: dfsub=pd.DataFrame()
                    else: dfsub=pd.read_excel(path+os.path.sep+filename,sheet_name=sheetname,skiprows=skiprows,na_values=['','NA','nan','null'],**read_excel_kwargs)             
            except Exception as e: 
                print('Cannot load:',filename,e)
                return None
            #keep leading zero when export to excel : df.A = df.A.apply('="{}"'.format)
            #xls = pd.ExcelFile(r"\\".join((path,filename))) #pd.ExcelFile(path+'\\'+filename).sheet_names
            if isinstance(dfsub,pd.DataFrame):
                if dfsub.empty: 
                    print('emty data of file',filename)
                    is_dataframe=True
            if not isinstance(dfsub,pd.DataFrame) and (is_setdf or len(filelist)==1 and isinstance(dfsub,dict) and len(dfsub)==1):
                #*** and not is_csv
                dfsub_=pd.DataFrame()
                for shtname in dfsub.keys(): dfsub_=dfsub_.append(dfsub[shtname],sort=False)
                dfsub=dfsub_
            if idx==0:
                is_dataframe=is_setdf or (isinstance(dfsub,pd.DataFrame) and (len(filelist)==1 or sheetname is not None))
                df=pd.DataFrame() if is_dataframe else collections.OrderedDict()
                #OrderedDict([('a','1'),('b','2')]), ddict=dict(OrderedDict) -->convert order dict to data dict         
            if is_dataframe:
                #dfsub=dfsub.loc[:,~dfsub.columns.astype(str).contains('^Unnamed')]
                dfsub=dfsub.loc[:,~dfsub.columns.str.contains('^Unnamed')] 
                for colname in dfsub.columns: #is_setdf to protect unwanted column (filename)
                    if 'unnamed' in str(colname).lower(): del dfsub[colname]     
                if not is_otm and (is_csv or is_setdf) and len(filelist)>0 and not dfsub.empty:
                    dfsub.loc[:,'filename']=os.path.splitext(filename)[0]
#                     dfsub.loc[:,'LastModifyTime']= datetime.datetime.fromtimestamp(os.stat(os.path.join(path,filename)).st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                df=df.append(dfsub,sort=False) 

            else : df.update({filename.replace('.csv',''): dfsub} if is_csv else dfsub) #df[filename]=dfsub

            print('Load '+str(len(dfsub))+(' rows' if is_dataframe else ' sheets')+' of : {} in {}'.format(filename,path))        
            sheetname=pre_sheetname        
        if len(filelist)>1:
            txt_filelistname=filelistname if type(filelistname)==str else ('[%s]' % ', '.join(map(str, filelistname)))
            print('==>Total ' + str(len(df))+(' rows' if is_dataframe else ' sheets')+' of : ' +txt_filelistname+'\n')
        if isinstance(df,pd.DataFrame): df=filter_columns(df.replace('\'','',regex=True))
    else: df=excelBlobCombine(path,filelistname)
    return df

def convert_xls_from_OTM(activedir,file_input=None,**read_excel_kwargs):
    import os
    import pandas as pd
    from os.path import join
    datadict = {} #datadict=collections.OrderedDict() #import collections
    fileinput=get_latest_file(activedir,file_input) #if (file_input is None or file_input=='') else file_input
    file_extension = os.path.splitext(fileinput)[1].lower()
    if not fileinput or file_extension not in ['.xls','.xlsx']:
        print('Cannot execute==>No input file or OTM download file extension must be "*.xls, *.xlsx"')
        return None
#     if file_extension=='.xls':
#         try: df=pd.read_csv(r"\\".join((activedir,fileinput)),sep='\t',low_memory=False,**read_excel_kwargs) #,error_bad_lines=False
#         except Exception as e: print('Error when tried to read file::',file_input,e)
#         df=df.loc[:, ~df.columns.str.contains('^Unnamed')]
#         write_sheets_to_file_from_datadict(activedir,fileinput.replace(file_extension,''), df)
#     else: df=load_data(activedir,fileinput,sheetname=None, skiprows=0,csv_sep=',',is_setdf=False,**read_excel_kwargs)
    df=load_data(activedir,fileinput,None,0,',',False,**read_excel_kwargs)
    if file_extension=='.xls': write_sheets_to_file_from_datadict(activedir,fileinput.replace(file_extension,''),df)
        #if 'filename' in df: df.drop(['filename'], axis=1, inplace=True)
        
    #if df is not None: print('---Read',len(df),'records---')
    if isinstance(df,pd.DataFrame): df.replace('\'','',inplace=True,regex=True)
    return df
#============================================================================================================
def savefile_by_key(active_dir,data_dict,key_list,set_index_list=None,add_filter_col=None):
    import re
    import os
    file_list=[]
    filter_key=key_list[0]
    exclude_in_add_filter=['%_Complete','count_ilm','count']
    if not os.path.isdir(active_dir):
        print('No folder to export ontime by transporter:',active_dir)
        return file_list
    trans_set=set()
    for key, df in data_dict.items():
        trans_set=trans_set.union(set(list(df[filter_key])))                                
    trans_set=[x for x in trans_set if x is not None]
    if len(trans_set)==0: return file_list                                
    
    print('Creating file and send mail by transporter...',trans_set)
    for x in trans_set:
        dict_out={}
        df_out=pd.DataFrame()
        df_=pd.DataFrame()
        for key, df in data_dict.items():
            df_=df[df[filter_key]==x]
            col_list=[m for m in df_.columns if m not in key_list]
            col_list2=[m for m in col_list if m not in exclude_in_add_filter]
            if len(df_)>0: 
                dict_out[key]=df_[col_list]
                df_out=df_out.append(df_,sort=False) #pd.concat([df_out,df_], axis=1, sort=False) #in case some df in dict has no data
                if add_filter_col is not None and not df_.empty: 
                    c1=df_[add_filter_col].isnull()|(df_['ILM_DOCUMENT'].fillna('')=='ไม่มีลายเซ็นลูกค้า')
                    if not df_[c1].empty: dict_out['NO_'+add_filter_col]=df_.loc[c1,col_list2]
                #if not df_.empty and add_filter_dict is not None and not df_[df_[next(iter(add_filter_dict.keys()))].isnull()].empty: 
                #  dict_out['NO_'+add_filter_dict]=df_.loc[df_[next(iter(add_filter_dict.keys()))].isnull(),next(iter(add_filter_dict.values()))]
        
        x=list(df_out[df_out[filter_key]==x][key_list].fillna('').apply(lambda x: '_'.join(map(str,x)),axis=1))[0]
        x=re.sub('[!@#$|.]', '', x)+'.xlsx'  #x.translate({ord(c): None for c in '!@#$|.'})      
        for key, df in dict_out.items():
            if set_index_list is not None and all(x in df.columns for x in set_index_list): 
                df.set_index(set_index_list,inplace=True)               
#         if add_filter_col is not None and not df_.empty:
#             c1=df_[add_filter_col].isnull()|(df_['ILM_DOCUMENT'].fillna('').str=='ไม่มีลายเซ็นลูกค้า')
#             if not df_[c1].empty: dict_out['NO_'+add_filter_col]=df_.loc[c1,col_list2]        
                
        #file_name=write_sheets_to_file_from_datadict(active_dir,x,dict_out,True)
        file_name=format_and_save(active_dir,x,dict_out)
        file_list.append(file_name)
    return file_list

def format_and_save(active_dir,output_name,dd,col=[0,0]):
    #Prepare File-------------------------------------------------------------------
    output_name=output_name.replace('.xlsx','').replace('.xls','')+ '_'+ datetime.datetime.strftime(datetime.date.today(),'%y%m%d')+'.xlsx'
    filename=active_dir+"\\"+output_name
    writer = pd.ExcelWriter(filename,engine='xlsxwriter')
    workbook=writer.book
    for key, df_ in dd.items():
        idx_columns=list(df_.index.names) #+list(df_.columns)
        if df_.empty: continue
        #print(output_name,key,idx_columns[0] is None,idx_columns,len(idx_columns))
        #SAVE FILE ----------------------------------------------------------------------------------------------
        df_.to_excel(writer,sheet_name=key,index=False)  # send df to writer
        worksheet=writer.sheets[key]  # pull worksheet object num_format='#,##0_);[Red](#,##0)'
        #COLUMN----------------------------------------------------------------------------------------------
        import string
        s=string.ascii_uppercase
        c_format=workbook.add_format({'valign':'vcenter'})
        worksheet.set_column(s[0]+':'+s[df_.index.nlevels-1],None,c_format)
        if idx_columns[0] is not None:
            for i in range(len(idx_columns)):
                max_len=max(df_.index.get_level_values(i).astype(str).map(len).max(),len(idx_columns[i]))+1
                worksheet.set_column(i,i,max_len)  # set column width
        col_list=list(df_.columns); idx_len=len(idx_columns)
        for i in range(len(idx_columns),idx_len+len(df_.columns)):
            max_len=max(df_.iloc[:,i-len(idx_columns)].astype(str).map(len).max(),len(col_list[i-idx_len]))+1
            worksheet.set_column(i,i,max_len)  # set column width, if Thai, vowels are count for width 
    #----------------------------------------------------------------------------------------------
    writer.save()
    writer.close()
    return output_name

def get_contacts(filename,csv_sep=' '):
    #print(ntpath.basename("a/b/c"),ntpath.split("a/b/c"),os.path.splitdrive("a/b/c"),os.path.split("a/b/c"))
    report=[]
    names=[]
    emails=[]
    extension=''
    
    if isinstance(filename,str): 
        if not os.path.isdir(os.path.split(filename)[0]): return None
        extension=os.path.splitext(filename)[1]
    else: df=filename
       
    if extension in ['','.csv','.xls','.xlsx']: 
        if extension!='':
            folder, file=os.path.split(filename)
            df=load_data(folder,file,None,0,',',False,dtype=str)
            if df is None: return None
        emails=list(df['email'])
        names=emails #list(df['name'])-->  names are displayed in sending list when open email
        if 'report' in df: report=list(df['report'].fillna(''))  
        else: 
            print('filter column must be"report"')
            return
    else:
        with open(filename, mode='r', encoding='utf-8') as contacts_file:
            #reader = csv.reader(file)
            #next(reader)  # Skip header row
            for a_contact in contacts_file:
                names.append(a_contact.split(csv_sep)[0])
                emails.append(a_contact.split(csv_sep)[1])              
    if len(emails)==0: return None
    z=zip(names,emails,report)
    return z
    
def read_template(filename):
    if not os.path.isfile(filename): return None
    from string import Template
    with open(filename, 'r', encoding='utf-8') as template_file:
        template_file_content = template_file.read()
    return Template(template_file_content)

def send_mail_yag(datadict,mail_config='mail',config_pathfile=config_pathfile): # First, need to Allow less secure apps: ON (https://myaccount.google.com/lesssecureapps)
    #import smtplib, ssl
    MailServerList=get_user_password(mail_config,config_pathfile)
    smtp_server=MailServerList[2] #"smtp.gmail.com" #'smtp-mail.outlook.com'
    MY_ADDRESS=MailServerList[0]  #sccc.logistics.excellence@gmail.com'
    PASSWORD =MailServerList[1]
    
    try: import yagmail #pip install yagmail
    except Exception as e: 
        txt='Fail: ***yagmail is not installed.***'
        print(txt)
        return txt
    
    contacts_pathfile=contacts_pathfile if 'contacts_file' not in datadict or isinstance(datadict['contacts_file'],str) and not os.path.isfile(datadict['contacts_file']) else datadict['contacts_file']
    if 'contacts_filter' not in datadict: datadict['contacts_filter']=''
    if 'msg' not in datadict: datadict['msg']=''
    if 'msg_html' not in datadict: datadict['msg_html']=''
    if 'msg_img' not in datadict: datadict['msg_img']=''
    if 'attached_file' not in datadict or datadict['attached_file']=='': datadict['attached_file']=None
    if 'msg_dict_key' not in datadict: 
        datadict['msg_dict_key']=[]
        datadict['msg_dict_value']=[]
    #-------------------------------------------------------------------  
    
    receiver_list=get_contacts(contacts_pathfile) # read contacts
    contacts_file_str=os.path.split(contacts_pathfile)[-1] if isinstance(contacts_pathfile,str) else 'df'
    contacts_filter=datadict['contacts_filter']
    contact_dict={} #****cannot print receiver_list before create contact_dict
    if receiver_list is not None:
        contact_dict={email:name for name,email,report in receiver_list if str(report).lower()==(str(report).lower() if contacts_filter=='' else contacts_filter.lower())} #if duplicated email, get the latest record
    
    if receiver_list is None or not contact_dict:
        txt='Fail: Mails not sent to "'+contacts_filter+'"('+contacts_file_str+') from '+MY_ADDRESS+'. Error:: no contacts'+ ('' if contact_dict else ' when filter')
        print(txt)
        return txt
#     for name, email in receiver_list:
#         if not isinstance(message,str): 
#             d[datadict['msg_dict_key'][0]]=name.title()
#             message=message.substitute(d)  #message_template.substitute(d) #message_template.safe_substitute()
#         yag.send(to={email:name} ,subject=datadict['subject'] ,contents=message ,attachments=datadict['attached_file'],) 
    #=========================================================================
    message=datadict['msg']  #message_template.substitute(d) #message_template.safe_substitute()
    if 'msg_dict_key' in datadict and datadict['msg_dict_key']:
        msg_dict_value_list=datadict['msg_dict_value']  #value name in message
        msg_dict_value_list=msg_dict_value_list+['']*(len(datadict['msg_dict_key'])-len(msg_dict_value_list))
    if not isinstance(message,str): message=message.substitute({x:y for x,y in zip(datadict['msg_dict_key'],msg_dict_value_list)})
    #message.safe_substitute({x:y for x,y in zip(datadict['msg_dict_key'],msg_dict_value_list)})
    #=========================================================================
    image_list=[yagmail.inline(x) for x in datadict['msg_img']] if isinstance(datadict['msg_img'],list) else [yagmail.inline(datadict['msg_img'])]
    msg_list=[message,datadict['msg_html']]+image_list
    msg_list=[x for x in msg_list if x is not None and x!='' ]
    yag=yagmail.SMTP(MY_ADDRESS,PASSWORD)
    try:
        yag.send(to=contact_dict ,subject=datadict['subject'] ,contents=msg_list ,attachments=datadict['attached_file'],)
        #send(self, to, subject, contents, attachments, cc, bcc, preview_only, headers, newline_to_break)
        isSend=True
        txt='Pass: '+str(len(contact_dict))+' mails sent to '+contacts_filter+'('+contacts_file_str+') from '+MY_ADDRESS
    except Exception as e:
        isSend=False
        txt='Fail: Mails not sent to "'+contacts_filter+'"('+contacts_file_str+') from '+MY_ADDRESS+'. Error::'+str(e)
    print(txt)
    return txt[:300]

def send_mail(datadict,mail_config='mail',config_pathfile=config_pathfile):
    import smtplib
    from email import encoders
    from email.mime.base import MIMEBase
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    
    is_ssl=True
    try: import ssl
    except Exception as e: 
        print('***ssl is error***')
        is_ssl=False
        
    userpwd_list=get_user_password(mail_config,config_pathfile)
    smtp_server=userpwd_list[2]
    MY_ADDRESS=userpwd_list[0]
    PASSWORD =userpwd_list[1]
    
    contacts_pathfile=contacts_pathfile if 'contacts_file' not in datadict or isinstance(datadict['contacts_file'],str) and not os.path.isfile(datadict['contacts_file']) else datadict['contacts_file']
    if 'contacts_filter' not in datadict: datadict['contacts_filter']=''
    if 'msg' not in datadict: datadict['msg']=''
    if 'msg_html' not in datadict: datadict['msg_html']=''
    if 'msg_img' not in datadict: datadict['msg_img']=''
    if 'attached_file' not in datadict or datadict['attached_file']=='': datadict['attached_file']=None
    if 'msg_dict_key' not in datadict: 
        datadict['msg_dict_key']=[]
        datadict['msg_dict_value']=[]
    #-------------------------------------------------------------------
    receiver_list=get_contacts(contacts_pathfile) # read contacts
    contacts_file_str=os.path.split(contacts_pathfile)[-1] if isinstance(contacts_pathfile,str) else 'df'
    contacts_filter=datadict['contacts_filter']
    #****cannot print receiver_list before create contact_dict 
    contact_dict={}
    if receiver_list is not None:
        contact_dict=[email for name,email,report in receiver_list if str(report).lower()==(str(report).lower() if contacts_filter=='' else contacts_filter.lower())] #if duplicated email, get the latest record
        
    if receiver_list is None or not contact_dict:
        txt='Fail: Mails not sent to "'+contacts_filter+'"('+contacts_file_str+') from '+MY_ADDRESS+'. Error:: no contacts'+ ('' if contact_dict else ' when filter')
        print(txt)
        return txt
    #contacts_filter=';'.join(contact_dict)
    #-------------------------------------------------------------------    
#     path,file=ntpath.split(contacts_pathfile)
#     contacts=load_data(path,file)
#     if contacts.empty: 
#         print('Mails not sent, no contacts.')
#         return
#     contact_dict =set(contacts.loc[contacts['report'].str.lower()==str(datadict['contacts_filter']).lower(),'email'])
#     if contact_dict is None or not contact_dict: 
#         print('Mails not sent, no contacts.')
#         return
#     contacts_filter=';'.join(contact_dict)
    #=========================================================================   
    message=datadict['msg']  #message_template.substitute(d) #message_template.safe_substitute()
    if 'msg_dict_key' in datadict and datadict['msg_dict_key']:
        msg_dict_value_list=datadict['msg_dict_value'] #value name in message
        msg_dict_value_list=msg_dict_value_list+['']*(len(datadict['msg_dict_key'])-len(msg_dict_value_list))
    if not isinstance(message,str): message=message.substitute({x:y for x,y in zip(datadict['msg_dict_key'],msg_dict_value_list)})
    #message.safe_substitute({x:y for x,y in zip(datadict['msg_dict_key'],msg_dict_value_list)})
    #=========================================================================
    msg=MIMEMultipart() #MIMEMultipart("alternative")   # create a message
    msg['Subject']=datadict['subject']
    msg['From']=MY_ADDRESS
    #msg['To']=contacts_filter #msg["Bcc"]=email  # Recommended for mass emails     
    msg.attach(MIMEText(message,'html')) #msg.attach(MIMEText(message, 'html'))
    #print(msg.as_string())
    #=========================================================================
    file_list=datadict['attached_file'] if isinstance(datadict['attached_file'],list) else [] if datadict['attached_file'] is None else [datadict['attached_file']] 
    for filename in file_list:
        part=None
        with open(filename,"rb") as attachment: # Open PDF file in binary mode
            # Add file as application/octet-stream
            # Email client can usually download this automatically as attachment
            part=MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
        if part is not None:
            encoders.encode_base64(part) # Encode file in ASCII characters to send by email    
            part.add_header("Content-Disposition",f"attachment; filename={os.path.split(filename)[-1]}",) # Add header as key/value pair to attachment part
            msg.attach(part) # Add attachment to message and convert message to string
    #=============================================================
    option=1
    try:## set up the SMTP server. First, need to Allow less secure apps: ON (https://myaccount.google.com/lesssecureapps)
        if option==1:        
            if is_ssl: context=ssl.create_default_context() #Create secure connection with server 
            s=smtplib.SMTP(host=smtp_server, port=587) #s=smtplib.SMTP_SSL(smtp_server, 465, context )
            if is_ssl: s.starttls(context=context) # For starttls
            else: s.starttls()
            s.login(MY_ADDRESS,PASSWORD)
        elif option==2:
            s=smtplib.SMTP(smtp_server, 587)
            s.ehlo()
            s.starttls()
            s.ehlo()
            s.login(MY_ADDRESS,PASSWORD)
        elif option==3: #SSLError: [SSL: WRONG_VERSION_NUMBER] wrong version number (_ssl.c:1056)
            if is_ssl: context=ssl.create_default_context() #Create secure connection with server 
            s=smtplib.SMTP(host=smtp_server, port=465) #s=smtplib.SMTP_SSL(smtp_server, 465, context )
            if is_ssl: s.starttls(context=context) # For starttls
            else: s.starttls()
            s.login(MY_ADDRESS,PASSWORD)
        elif option==4: #TimeoutError: [WinError 10060] A connection attempt failed because the connected party did not properly respond after a period of time
            context=ssl.create_default_context()
            with smtplib.SMTP_SSL(smtp_server, 465, context=context) as s:
                s.login(MY_ADDRESS,PASSWORD)
        #s.send_message(msg)
        s.sendmail(MY_ADDRESS,contact_dict,msg.as_string()) #s.sendmail(MY_ADDRESS,email,message.format(PERSON_NAME=name,GRADE=grade),)
        s.quit() # Terminate the SMTP session and close the connection
        del msg
        isSend=True
        txt='Pass: '+str(len(contact_dict))+' mails sent to '+contacts_filter+'('+contacts_file_str+') from '+MY_ADDRESS
    except Exception as e:
        isSend=False
        txt='Fail: Mails not sent to "'+contacts_filter+'"('+contacts_file_str+') from '+MY_ADDRESS+'. Error::'+str(e)
    print(txt)
    return txt[:300]
        
def send_mail_outlook(datadict): #need to install outlook
    from win32com.client.dynamic import Dispatch #Dispatch #DispatchEx # from win32com.client.gencache import EnsureDispatch
    import ntpath
    outlook=Dispatch('Outlook.Application')
    mail=outlook.CreateItem(0)
    
    contacts_pathfile=contacts_pathfile if 'contacts_file' not in datadict or isinstance(datadict['contacts_file'],str) and not os.path.isfile(datadict['contacts_file']) else datadict['contacts_file']
    path,file=ntpath.split(contacts_pathfile)
    contacts=load_data(path,file)
    if contacts.empty: return
    contacts_filter=contacts[contacts['report'].str.lower()==datadict['contacts_filter']]
    if contacts_filter is None: return
    if 'msg_html' not in datadict: datadict['msg_html']=''
    if 'msg_img' not in datadict: datadict['msg_img']=''
    if 'attached_file' not in datadict or datadict['attached_file']=='': datadict['attached_file']=None
    
    mail.Subject=datadict['subject']
    mail.To =';'.join(set(contacts_filter['email'])) #mail.CC = 'teeradaj.tuntiwisas@siamcitycement.com'
    message=datadict['msg']  #message_template.substitute(d) #message_template.safe_substitute()
    msg_dict_value_list=datadict['msg_dict_value'] if datadict['msg_dict_value'] is not None else []  #value name in message
    msg_dict_value_list=msg_dict_value_list+['']*(len(datadict['msg_dict_key'])-len(msg_dict_value_list))
    if not isinstance(message,str): message=message.substitute({x:y for x,y in zip(datadict['msg_dict_key'],msg_dict_value_list)})
    mail.HTMLBody=message+datadict['msg_html']  #wb=excel.Workbooks.Open(excel_path) ws=wb.Worksheets(1) ws.Range("A1:B2").Copy() wb.Close()
    #mail.Body=message
    #image_list=[x for x in image_list if os.path.isfile(x)]
    image_list=datadict['msg_img'] if isinstance(datadict['msg_img'],list) else [] if datadict['msg_img'] is None or datadict['msg_img']=='' else [datadict['msg_img']]
    for y in image_list:
        mail.Attachments.Add(y) 
        
    file_list=datadict['attached_file'] if isinstance(datadict['attached_file'],list) else [] if datadict['attached_file'] is None else [datadict['attached_file']]
    for x in file_list:
        mail.Attachments.Add(x) 
    
    try: 
        mail.Send()
        print('Mail sent')
    except Exception as e: print('Cannot sent mail, check whether Outlook is opened.','Error::',e)
#             #Set Table ------------------------------------------ 
#             from tabulate import tabulate #pip install tabulate
#             table=tabulate(df, headers=df.columns, tablefmt="grid") #headers="firstrow"
#             table=df.to_html(index=True,border=1).replace('<tr>', '<tr align="center">') #open(myfile, 'w').write(html)
#             #------------------------------------------
#             #COPY from MS WOrd then PASTE in Outlook------------------------------------------
#             word=EnsureDispatch('Word.Application')
#             doc=word.Documents.Add() #doc=word.Documents.Open(result_folder+r'\doc.docx')
#             doc.Content.PasteExcelTable(False, False, True)
#             doc.Content.Copy()
#             doc.Close(SaveChanges=False)
#             #-----------------
#             mail.BodyFormat =2
#             mail.GetInspector.WordEditor.Range(Start=0, End=0).Paste()
#             mail.Display(False) #need, unless no table in mail
#             ##word.Application.Quit() #SaveChanges=False
#             #------------------------------------------
#============================================================================================================
def import_pywin32(): #http://www.xavierdupre.fr/blog/2014-07-01_nojs.html
    """
    For the module ``pywin32``,
    this function tries to add the path to the DLL to ``PATH``
    before throwing the exception: 
    ``DLL load failed: The specified module could not be found``.
    """
    try:
        import win32com
    #except ImportError as e :
    except Exception as e :
        if "DLL load failed:" in str(e):
            import os,sys
            path = os.path.join(os.path.split(sys.executable)[0], "Lib","site-packages","pywin32_system32")
            os.environ["PATH"] = os.environ["PATH"] + ";" + path
            print('import pywin32',path)
            try:
                import win32com
            except ImportError as ee :
                dll = os.listdir(path)
                dll = [os.path.join(path,_) for _ in dll if "dll" in _]
                raise ImportError("some DLL must be copied:\n" + "\n".join(dll)) from e
        else :
            raise e
            
def copy_sheet(path1,path2,option=1):
    # path1=(path1.replace('\\\\','\\'))
    # path2=(path2.replace('\\\\','\\'))
    file1=os.path.split(path1)[-1]
    file2=os.path.split(path2)[-1]
    error=''
    if option==1:
        ##---------------------------------------------
        #https://stackoverflow.com/questions/47608506/issue-in-using-win32com-to-access-excel-file/47612742
        #https://gist.github.com/rdapaz/63590adb94a46039ca4a10994dff9dbe
        # If errors are found, do this
        # clear contents of C:\Users\<username>\AppData\Local\Temp\gen_py
        # that should fix it, then test
        from win32com.client.gencache import EnsureDispatch #***trigger by this, to ensure ".Copy(After)" to work
        xl=EnsureDispatch('Excel.Application')
        ##---------------------------------------------
#         from win32com.client.dynamic import Dispatch #DispatchEx # from win32com.client.gencache import EnsureDispatch
#         xl=Dispatch("Excel.Application")
        ##---------------------------------------------
        #from win32com.client import Dispatch #instead of  win32com.client.gencache.EnsureDispatch
        #xl=Dispatch("Excel.Application")  #not work with .Copy(After) for some computer
        ##---------------------------------------------
        xl.Visible=False
        xl.DisplayAlerts=False
        #xl.ScreenUpdating=False
        #xl.EnableEvents=False
        try: 
            wb1=xl.Workbooks.Open(Filename=os.path.abspath(path1)) #if not abspath, cannot save in oneDrive
            wb2=xl.Workbooks.Open(Filename=os.path.abspath(path2))
            ws2_count=wb2.Worksheets.Count #wb2.Sheets.Count
            for x in range(wb1.Sheets.Count):
                wb1.Worksheets(x+1).Copy(After=wb2.Worksheets(ws2_count+x))
                # wb1.Worksheets(x+1).Copy(Before=wb2.Worksheets(1))
                #*** .Copy(After) is not work, need to trigger by xl=EnsureDispatch('Excel.Application') first
            wb1.Close(SaveChanges=False)
            #wb2.BreakLink(Name=r"C:\Users\me\dummy.xlsx", Type=1)
            print('Copy sheet and save to file:',os.path.abspath(path2),'....')
            # wb2.SaveAs(os.path.abspath(path2))
            wb2.Save() #***works good for saving in Onedrive
            wb2.Close() #SaveChanges=True --> still ask for saving, use wb2.SaveAs(path2)
            print('Copy complete: from',file1,'to',file2)
        except Exception as e:
            error="**** Copy fail: from "+file1+' to '+file2+'::'+str(e)
            #::**** Copy fail: from 02_DailyTrackingNonCement_Feb -2021.xlsx to dl_tracking_2102-Feb_210204.xlsx::(-2147352567, 'Exception occurred.', (0, 'Microsoft Excel', "Cannot access 'dl_tracking_2102-Feb_210204.xlsx'.", 'xlmain11.chm', 0, -2146827284), None)
            print(error)
            #sys.exit(1)
        finally:
            xl.Quit()
            del xl 
    if option==2:
        import xlwings as xw
        xw.Visible=False
        xw.DisplayAlerts=False
        try:  
            wb1=xw.Book(path1)
            wb2=xw.Book(path2)
            ws2_count=wb2.sheets.count #wb2.sheets.count       
            for x in range(wb1.sheets.count):
                wb1.sheets(x+1).api.Copy(After=wb2.sheets(ws2_count+x).api) #Before=wb2.sheets(1)
                #*** .Copy(After) is not work, need to trigger by xl=EnsureDispatch('Excel.Application') first
            wb2.save(path2)
            print('Copy complete: from',file1,'to',file2)
        except Exception as e:
            error="**** Copy fail: from "+file1+' to '+file2+'::'+str(e)
            print(error)
            #sys.exit(1)
        finally: 
            wb1.app.quit()
            wb2.app.quit()
    time.sleep(5)
    return error
        
def save_picture(filename,dd=None,default_name='',prefix=''): #prefix=transporter code
    #Table html------------------------------------------
    #import_pywin32()
    from win32com.client.dynamic import Dispatch #Dispatch #DispatchEx # from win32com.client.gencache import EnsureDispatch
    from win32com.client import constants  
    from PIL import ImageGrab
    import string
    import ntpath
    activedir,file=ntpath.split(filename)
    is_fromfile=False ;is_1sheet=False ;error=''
    #-------------------
    if dd is None:
        is_fromfile=True
        picname=default_name if default_name else '_'
        df=load_data(activedir,file)
        if df is None: 
            error='No data from '+filename
            return [],error
        elif isinstance(df,pd.DataFrame):
            is_1sheet=True
            dd={picname:df}
        else: dd=df
    #htmfilename=filename.replace('xlsx','htm')
    xl=Dispatch('Excel.Application')
    xl.Visible=False
    xl.DisplayAlerts=False
#     xl.ScreenUpdating=False cannot use
    xl.EnableEvents=False
    try:
        wb=xl.Workbooks.Open(filename)
        #wb.SaveAs(htmfilename,constants.xlHtml)
        #-------------------   
    #         table=open(htmfilename).read(), table=html_txt.format(df.to_html())
    #         table=html_txt.format(table=table) #This pag '\n<br><br>'+e uses frames, but your browser doesn't support them.
        attach_list=[]
        for key, df in dd.items():
            if df.empty: continue
            sheetname=wb.Worksheets(1).Name if is_1sheet else key
            s=string.ascii_uppercase
            s_=['','A','B','C']
            s=[x+y for x in s_ for y in s]
            #if reload from file , no need to use nlevels
            #print(key,df.columns.nlevels,len(df.columns),len(df.index),'\n',df.columns,df.index)
            #print(df.columns.nlevels,len(df.columns),len(df.index))
            if is_fromfile: data_range="A1:"+s[len(df.columns)-1]+str(len(df.index)+df.columns.nlevels)
            else: data_range="A1:"+s[len(df.columns)+df.index.nlevels-1]+str(len(df.index)+df.columns.nlevels+1)
            image_name=prefix+sheetname+'.png'
            image_path=activedir+r'\\'+image_name    
            print('copy:',sheetname,data_range,'to',image_name)
            ws=wb.Worksheets(sheetname) #wb.Worksheets(1).Name ,wb.Sheets(1).Name
            try:
                ws.Range(data_range).CopyPicture(Format=2)#(Format=constants.xlBitmap) #ws.Range(data_range).Copy()   #ws.Range("A1:K5").Copy(ws.Range("A7:K11"))
                img=ImageGrab.grabclipboard()
                print('Copied sheet',sheetname,'to picture:',os.path.abspath(image_path),'....')
                img.save(os.path.abspath(image_path)) #,'png'
                attach_list.append(image_name)
                #image_list.extend([key,image_name])
                time.sleep(1)
            except Exception as e: 
                error+=' NOT save '+image_name+', AttributeError::'+str(e)
                continue
            #htmfilename= active_dir+r'\dl_tracking\test_files\sheet001.htm'  
        #--------------------------------
        xl.Workbooks.Close()
        xl.Quit()
        del xl
        
    except Exception as e: error+=str(e) #('Cannot save '+image_name+', please re-run report. AttributeError::',e, file=sys.stderr)
    if error!='': print(error+' please re-run report')    
    return attach_list,error

def concat_images(image_list,ver_or_hor,active_dir=None):
    from PIL import Image
    #image_list=[os.path.join(active_dir,x) for x in image_list]
    error=''
    if active_dir is not None: 
        active_dir=active_dir+os.path.sep if os.path.splitdrive(active_dir)[0]==active_dir else active_dir
        image_list=[os.path.join(active_dir,x) for x in image_list]
    image_list=[x for x in image_list if os.path.isfile(x)]
    if not image_list: return '','No image list'+str(image_list)
    
    active_dir=os.path.split(image_list[0])[0]
    extension=os.path.splitext(os.path.split(image_list[0])[1])[-1]
    image_file_list=[os.path.splitext(os.path.split(x)[-1])[0] for x in image_list]
    new_pathfile=''
    try: 
        images = [Image.open(x) for x in image_list]
        widths, heights = zip(*(i.size for i in images))
        gap=10
        total_width = max(widths) if ver_or_hor=='v' else (gap*(len(images)-1)+sum(widths))
        total_height= (gap*(len(images)-1)+sum(heights)) if ver_or_hor=='v' else max(heights)
        new_im = Image.new('RGB', (total_width, total_height))

        x_offset=0 ;y_offset=0
        for im in images:
            new_im.paste(im, (x_offset,y_offset)) #0 is vertical offset
            if ver_or_hor=='h': x_offset += (im.size[0]+gap)
            else: y_offset += (im.size[1]+gap)
    
        new_pathfile=active_dir+r'\\'+'_'.join([x.split('_')[0] for x in image_file_list])+extension
        new_im.save(new_pathfile)
    except Exception as e:
        error=str(e)
        print(error) #e.message, e.args
        
    return new_pathfile,error

def compare_file(var,df=None,isCompare=False):
#     [{'active_dir':active_dir+r'\\'+FreightCal_sheet+r'\\'+yearmonth
#       ,'rename_dict':{'Shipment ID':'ShipmentID','Pre DO Number':'PreDONo','Final Freight':'FinalFreightCost','filename':'FinalFreightFile'}
#       ,'col_list':['ShipmentID','PreDONo','FinalFreightCost','FinalFreightFile']
#       ,'key':['ShipmentID','PreDONo']}]
    import re
    filelistname='*.xls*'
    active_dir=var['active_dir']
    rename_dict=var['rename_dict']
    col_list=var['col_list']
    key_list=var['key']
    dd_dict={}
    df_data=None
    if df is not None:
        #df=df.reindex(columns=list(set(df.columns).union(set(col_list))))
        for x in col_list:
            if x not in df.columns: df.loc[:,x]=None
    if isCompare and os.path.isdir(active_dir):
        datadict={x:str for x in ['Delivery Order No','PreDONo']}
        #df_data=load_data(active_dir,filelistname,None,0,',',True,converters=datadict)#dtype=str
        df_data=get_df_from_df_latest_file(active_dir,'',converters=datadict)
        if df_data is not None:
            df_data.rename(index=str,columns=rename_dict,inplace=True) #[col_list]
            df_data=df_data.dropna(subset=key_list) #df_data[df_data[col_list[0]].notnull()]
            for x in key_list:
                df_data[x].replace('ILOG.','',regex=True,inplace=True)

            name=re.sub(r'[!@#$|.:*?/\\]', '', '_'+active_dir[active_dir.rfind('_')+1:])[:10]
            write_sheets_to_file_from_datadict(active_dir,name,df_data,True)
            df_data=df_data[[x for x in col_list if x in df_data.columns]]
        if df is not None and df_data is not None and len(df_data)>0:
            for x in set(col_list)-set(key_list):
                if x in df.columns: del df[x] #*** delete columns
            df=df.copy().merge(df_data, how='left') 
            #cols=set(df.columns)-(set(col_list)-set(key_list)) #df.columns.difference(df_data.columns)
            #df=df[cols].copy().merge(df_data,how='left')
    else: print('Not compare or No folder of',active_dir)
    #w['female'] = w['female'].map({'female': 1, 'male': 0})
    dd_dict['df']=df
    dd_dict['df_load']=df_data
    #if df is not None and isCompare and df_data is not None and len(df_data)>0: print('Join to :',str(len(df[df[col_list[-1]].notnull()])),'rows in',active_dir)
    return dd_dict

def copy_tree(src, dst, preserve_mode=1, preserve_times=1,
              preserve_symlinks=0, update=0, verbose=1, dry_run=0):
    """Copy an entire directory tree 'src' to a new location 'dst'.
    Both 'src' and 'dst' must be directory names.  If 'src' is not a
    directory, raise DistutilsFileError.  If 'dst' does not exist, it is
    created with 'mkpath()'.  The end result of the copy is that every
    file in 'src' is copied to 'dst', and directories under 'src' are
    recursively copied to 'dst'.  Return the list of files that were
    copied or might have been copied, using their output name.  The
    return value is unaffected by 'update' or 'dry_run': it is simply
    the list of all files under 'src', with the names changed to be
    under 'dst'.
    'preserve_mode' and 'preserve_times' are the same as for
    'copy_file'; note that they only apply to regular files, not to
    directories.  If 'preserve_symlinks' is true, symlinks will be
    copied as symlinks (on platforms that support them!); otherwise
    (the default), the destination of the symlink will be copied.
    'update' and 'verbose' are the same as for 'copy_file'.
    """
    from distutils.dir_util import mkpath
    from distutils.file_util import copy_file
    error=''
    if not dry_run and not os.path.isdir(src):
        error="error: cannot copy tree '%s': not a directory" % src
        raise DistutilsFileError(error)
        return error
    try:
        names = os.listdir(src)
    except os.error as e:
        if dry_run:
            names = []
        else:
            error="error listing files in '%s': %s" % (src, str(e))
            raise DistutilsFileError(error)
            return error
    outputs = []
    try:
        if not dry_run: mkpath(dst, verbose=verbose)
        for n in names:
            src_name = os.path.join(src, n)
            dst_name = os.path.join(dst, n)
            if n.startswith('.nfs'):
                # skip NFS rename files
                continue
            if preserve_symlinks and os.path.islink(src_name):
                link_dest = os.readlink(src_name)
                if verbose >= 1:
                    log.info("linking %s -> %s", dst_name, link_dest)
                if not dry_run:
                    os.symlink(link_dest, dst_name)
                outputs.append(dst_name)
            elif os.path.isdir(src_name):
                outputs.extend(
                    copy_tree(src_name, dst_name, preserve_mode,
                              preserve_times, preserve_symlinks, update,
                              verbose=verbose, dry_run=dry_run))
            else:
                copy_file(src_name, dst_name, preserve_mode,
                          preserve_times, update, verbose=verbose,dry_run=dry_run)
                outputs.append(dst_name)
        print('copy from:',src,' to:',dst)
        return outputs
    except Exception as e:
        error='error: cannot copy file to {}::{}'.format(dst,e)
        return error

def revise_columns_with_master(df,script,var_dict,col_list,dbname=defaultdb):
    var_list4=store_var(script)
    for key, var in var_dict.items():
        var_list4[key]=var #'Master_Tzone_SAP'
    query=store_query(script) # "select * from ShipmentTracking where shipmentID like 'SH20180801%'"
    df_=sql_execute_query(dbname,query,var_list4,False)
    #any(x in y for x in i for y in a)
    mapcol=[x for x in df_.columns if x not in col_list]
    for x in mapcol:
        if x in df.columns: df.loc[:,x]=df[x].astype(str).replace(['nan'],[None],regex=True)       
    col_dict={x:x+'_' for x in col_list if x in df.columns}
    df_.rename(columns=col_dict,inplace=True)
    df=df.merge(df_, how='left')
    for x in col_list:
        if x+'_' in df.columns:
            df.loc[:,x]=df[x+'_'].fillna(df[x])
            del df[x+'_']
    print('Finish revise_columns_with_master()')
    return df

def to_time_delta(number):
    import math
    return pd.to_timedelta(f'{math.floor(number)}hours {(number - math.floor(number)) * 60}min')

def mask(df, key, value):
    if value is None or value==None: return df[df[key].isnull()], df[df[key].notnull()]
    return df[df[key]==value], df[df[key]!=value]

def merge_df(varlist_df_dict):
    if isinstance(varlist_df_dict,dict) and varlist_df_dict: #Empty dictionaries evaluate to False
        key_list=varlist_df_dict['key_list']
        df1=varlist_df_dict['df']
        df2=varlist_df_dict['df2']
        replace_list=varlist_df_dict['replace_list'] if any('replace_list' in d for d in varlist_df_dict) else []
        if df1.empty or df1 is None: return df1
        if replace_list: df1=df1.loc[:,[x for x in df1.columns if (x not in replace_list) or (x in key_list)]]
        df2=df2.loc[:,[x for x in df2.columns if (x not in df1.columns) or (x in key_list)]]
        df=df1.merge(df2,how='left',on=key_list)
        return df
    else: return varlist_df_dict

#LINE===========================================
def linenotify(group='',key='',dbname=defaultdb): #if df_token is None: df_token=t.linenotify(dbname=t.defaultdb) #both None&'' in send_mail()
    if dbname is None: return linenotify(group,key,linenotify())
    df=None ;is_dbname=isinstance(dbname,str) and dbname #check isinstance first
    if is_dbname:
        var_list2=store_var('search_master')
        var_list2['table']='master_token'
        var_list2['doc_no']=group
        var_list2['doc_name']=key
        #df,var_list2=get_df_from_sql('search_master',var_list2)   
        query=store_query('search_master')
        df=sql_execute_query(dbname,query,var_list2,False)
    elif isinstance(dbname,pd.DataFrame): df=dbname
        
    if df is None or df.empty or is_dbname and not(group or key): 
        print('Line token is {}'.format('NONE' if df is None or df.empty else '{} records'.format(len(df))))
        return df
    if group=='tea' or key=='tea': df=df.loc[df['group']=='tea',:]
    #elif group=='' and key=='': pass --> ***do not use all
    elif group and key: df=df.loc[(df['group'].fillna('').str.lower()==group.lower())&(df['key'].fillna('').str.lower()==key.lower()),:]
    elif key: df=df.loc[df['key'].fillna('').str.lower()==key.lower(),:]
    elif group: df=df.loc[df['group'].fillna('').str.lower()==group.lower(),:]
    else: df=pd.DataFrame()
    print('Line token filtered of group={}, key={} : {} records'.format(group,key,len(df)))
    return df

def notifyText(token,message):
    payload = {'message':message}
    return _lineNotify(token,payload)

def notifyFile(token,filename,message=' '):
    payload = {'message':message}
    return _lineNotify(token,payload,filename)

def notifyPicture(token,url):
    payload = {'message':" ",'imageThumbnail':url,'imageFullsize':url}
    return _lineNotify(token,payload)

def notifySticker(token,stickerPackageID,stickerID):
    payload = {'message':" ",'stickerPackageId':stickerPackageID,'stickerId':stickerID}
    return _lineNotify(token,payload)

def _lineNotify(token,payload,filename=''):
    import requests
    url = 'https://notify-api.line.me/api/notify'
#     if (isinstance(token,str) and not token) or token.empty: 
#         print('Token is empty')
#         return False
    if isinstance(token,str):
        headers={'Authorization':'Bearer '+token} #'content-type':'application/x-www-form-urlencoded',
    #     session=requests.Session()
    #     r=session.post(url, headers=LINE_HEADERS, files=file, data=data)
        try:
            file={'imageFile':open(filename,'rb')} if filename else None
            response=requests.post(url, headers=headers ,data=payload ,files=file)
            print(str(response.status_code)+'_'+payload['message'])
            return True
        except Exception as e:
            print('Line not sent::',e) #e.message, e.args
            return False
    else:
        line_dict={}
        try:
            for idx,tok in token.iterrows():
                headers={'Authorization':'Bearer '+str(tok['token'])}#'content-type':'application/x-www-form-urlencoded',
                file={'imageFile':open(filename,'rb')} if filename else None #must open everytimes unloess error 400
                response=requests.post(url, headers=headers ,data=payload ,files=file)
                line_dict.update({(tok['group'] if pd.isna(tok['key']) else tok['key']):str(response.status_code)+'_'+payload['message']}) #200: OK
            if line_dict: print('Line sent::',line_dict)
            return True
        except Exception as e:
            print('Line not sent::',e) #e.message, e.args
            return False

#SELENIUM===========================================
#https://medium.com/@sangamsyabil/an-effective-way-to-download-file-s-selenium-webdriver-53e55b92e08e
def otm_query(path_file,var=None):
    import re
    if os.path.isfile(path_file): text=read_template(path_file).safe_substitute()
    else: text='Nofile: {}'.format(path_file)
    rev_list=[]
    if var and isinstance(var,list):
        rev_list=[{'from':">=TO_DATE\('.*'YYYY-MM-DD'\)" ,'to':">=TO_DATE('"+var[0]+"','YYYY-MM-DD')"}
                  ,{'from':"<=TO_DATE\('.* 23:59:59" ,'to':"<=TO_DATE('"+var[1]}
                 ]
        if 'shipment_missingdo' in path_file.lower(): 
            days=0 ;min_time=datetime.datetime.strptime('08:29:00','%H:%M:%S').time() ;max_time=datetime.datetime.strptime('16:59:00','%H:%M:%S').time()
            if datetime.datetime.now().time()<=min_time or datetime.datetime.now().time()>=max_time: days=12 
            else: days=1
            rev_list.extend([{'from':"INSERT_DATE>=TRUNC\(SYSDATE\)" ,'to': f"INSERT_DATE>=TRUNC(SYSDATE)-{days}--"}])
    elif var and isinstance(var,dict): rev_list=[var]
    elif isinstance(var,str):
        #text+="\n AND SHIPMENT_XID IN ("+var+")"
        rev_list=   [{'from':"AND \(SH.ATTRIBUTE_DATE2" ,'to':"--AND \(SH.ATTRIBUTE_DATE2"}
                    ,{'from':"OR ORL.ATTRIBUTE_DATE2"  ,'to':"--OR ORL.ATTRIBUTE_DATE2"}
                    ,{'from':"OR ORL.ATTRIBUTE_DATE8"  ,'to':"--OR ORL.ATTRIBUTE_DATE8"}
#                   ,{'from':"OR SH.INSERT_DATE"
#                     ,'to':"--OR SH.INSERT_DATE"}
#                   ,{'from':"AND GLOG_UTIL.REMOVE_DOMAIN\(SS.STATUS_VALUE_GID\)"
#                     ,'to':"--AND GLOG_UTIL.REMOVE_DOMAIN(SS.STATUS_VALUE_GID)"}
                    ,{'from':"AND SHIPMENT_XID" ,'to':"--AND SHIPMENT_XID"}
                    # COST BY DO -------------------------
                    ,{'from':"AND ORL.ATTRIBUTE1",'to':"--"}
                    ,{'from':"AND \(ORL.ATTRIBUTE_DATE2",'to':"--"}
                    ,{'from':"OR ORL.ATTRIBUTE_DATE4",'to':"--"}
                    ,{'from':"OR OM.INSERT_DATE",'to':"--"}
                    #Shipment Report ,COST BY DO ------------------------------------------
                    ,{'from':"1=1" ,'to':var}
                    #Distance By Stop ------------------------------------------
                    ,{'from':"SH.ATTRIBUTE1 NOT IN" ,'to':var+' AND SH.ATTRIBUTE1 NOT IN'}
                    ]
    for x in rev_list:
        text=re.sub(x['from'],x['to'],text)
    return text

def every_downloads_chrome(driver):
    #driver.switch_to.window(driver.window_handles[-1]) # always switch to new tab
    if not driver.current_url.startswith("chrome://downloads"): driver.get('chrome://downloads') # navigate to chrome downloads to refresh browser every x secs
    #if not driver.current_url.startswith("chrome://downloads"): driver.get("chrome://downloads/") # to not refresh browser
    return driver.execute_script("""var items = downloads.Manager.get().items_;
                                    if (items.every(e => e.state === "COMPLETE"))
                                        return items.map(e => e.fileUrl || e.file_url);""")

# [state, percent, progress] = driver.execute_script("""
#     var item = downloads.Manager.get().items_[0];
#     var state = item.state;
#     var percent = item.percent;
#     var progress = item.progressStatusText;
#     return [state, percent, progress];
#     """)

def getDownLoadedFileName(driver,waitTime):
    #driver.switch_to.window(driver.window_handles[-1]) # always switch to new tab
    driver.get('chrome://downloads') # navigate to chrome downloads
    endTime = time.time()+waitTime # define the endTime
    while True:
        try:# get downloaded percentage
            downloadPercentage = driver.execute_script(
                "return document.querySelector('downloads-manager').shadowRoot.querySelector('#downloadsList downloads-item').shadowRoot.querySelector('#progress').value")
            # check if downloadPercentage is 100 (otherwise the script will keep waiting)
            if downloadPercentage == 100: # return the file name once the download is completed
                return driver.execute_script("return document.querySelector('downloads-manager').shadowRoot.querySelector('#downloadsList downloads-item').shadowRoot.querySelector('div#content  #file-link').text")
        except Exception as e: 
            print('fail getting downloadPercentage',e)
            pass
        time.sleep(3)
        if time.time() > endTime: break

def newChromeBrowser(headless=True, downloadPath=None, chromePath=None): #**downloadPath
    from webdriver_manager.chrome import ChromeDriverManager
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    """ Helper function that creates a new Selenium browser """
    downloadPath=downloadPath.replace('\\\\\\',os.sep).replace('\\\\',os.sep).replace('/',os.sep).replace('\script\..','')
    options=webdriver.ChromeOptions()
    # options=Options()
    if headless: options.add_argument('headless')
    # options.add_argument("--headless")
    # options.add_argument("--window-size=1920x1080")
    if downloadPath is not None:
        prefs = {} #os.makedirs(downloadPath)
        prefs["profile.default_content_settings.popups"]=0
        prefs["download.default_directory"]=downloadPath
        prefs["download.prompt_for_download"]= False #revised 9/6/21 from "false" #option
        prefs["safebrowsing.enabled"]=True #revised 9/6/21 from "false"  #option  
        prefs["download.directory_upgrade"]=True #added 9/6/21 from  #option
        options.add_experimental_option("prefs", prefs)
    chromePath=(downloadPath if chromePath is None else chromePath)+r"\chromedriver.exe"
    print('Download path',downloadPath,'(***must has single slash "\\") \nChrome path:',chromePath)
    #browser = webdriver.Chrome(options=options, executable_path=chromePath)
    browser = webdriver.Chrome(ChromeDriverManager().install(),options=options)
    return browser #driver

def get_download_filename(activedir):
    before=after = os.listdir(activedir)
    file_name='crdownload'
    while set(after) == set(before) or 'crdownload' in file_name:
        time.sleep(3)
        after = os.listdir(activedir)
        change = set(after) - set(before)
        if len(change) == 1:
            file_name = change.pop()
    return file_name

def downloads_done(activedir,option=1,driver=None):
#     for i in os.listdir(activedir):
#         if "crdownload" in i:
#             time.sleep(4)
#             downloads_done()
#     return get_latest_file(activedir)
#----------------------------------------------
    from selenium.webdriver.support.ui import WebDriverWait 
    WaitTime=480
    file_name=''
    #time.sleep(1)
    if not os.path.exists(activedir) or not 'tmp' in activedir:
        print(f'No temp dir to download; {activedir}')
        return file_name
    if option==1: #download to temp dir first 
        file_name='crdownload' 
        before=after=os.listdir(activedir) ;start=end=time.time()
        print(f'Start download to: {activedir}')
        #ex. Unconfirmed 61825.crdownload ,9cd208ca-1791-4339-a744-9a256c41ca5c.tmp
        while (set(after)==set(before) or any(word in file_name for word in ['crdownload','.tmp'])) and end-start<=WaitTime: #4 minutes
            time.sleep(2)
            after=os.listdir(activedir)
            change=set(after)-set(before)
            end = time.time()
            # diff = end-start ;print(f'End-Start:{diff} ,{activedir} ,{file_name}')
            if len(change)==1:
                file_name = change.pop() # removes and returns last value from the list
                if not 'crdownload' in file_name and ('.xls' in file_name or '.' not in file_name): break
        print(f'End download {file_name}:',round(end-start),'secs.')
        if file_name: #'crdownload' not in file_name and ('.xls' in file_name or '.' not in file_name):
            path_file=ntpath.split(activedir)[0] +r'\\'+ file_name
            shutil.copyfile(activedir +r'\\'+ file_name ,path_file)
            if os.path.isfile(path_file):
                time.sleep(3)
                shutil.rmtree(activedir)
                print('Move download to:',path_file)
            else: 
                print(f'Cannot move download file from tmp {activedir} to {path_file}')
                file_name='' #to throw error
        #else: file_name=''  
    if option==2:
        #This is no more work after chrom version 80 (2020 feb 06 )
        #https://gist.github.com/ic0n/a38b354cac213e5aa50c55a0d8b87a0b
        varURL=''
        driver.execute_script("window.open()") #***must open new tab after download tab
        driver.switch_to.window(driver.window_handles[-1]) # switch to new tab
        varURL=WebDriverWait(driver,WaitTime,1).until(every_downloads_chrome(driver))[0] #refresh every 3 secs
        if varURL is not None and varURL!='': 
            driver.quit()
    #     varFixedPath=getDownLoadedFileName(driver,20) #waiting 3 minutes to complete the download
    #     error varURL'file:///D:/Siam%20City%20Cement%20Public%20Company%20Limited/Logistics_Reports%20-%20Documents/xxx.xls' fix below
    #     import pathlib
    #     varFilename = sys.argv[0]
    #     varFilename = os.path.abspath(varFilename)
    #     varMainFolder = os.path.dirname(varFilename)
    #     varTarget = os.path.join(varMainFolder,"test test.py")
    #     varURL = pathlib.Path(varTarget).as_uri()
    #     varURL='file:///D:/Siam%20City%20Cement%20Public%20Company%20Limited/Logistics_Reports%20-%20Documents/1572435506523.xls'
            import urllib.parse
            varPathRaw = urllib.parse.urlparse(varURL).path
            varPathDecode = urllib.parse.unquote(varPathRaw)
            varOSPath = os.path.normpath(varPathDecode)
            varFixedPath = varOSPath
            varDrive = os.path.splitdrive(varFixedPath[1:])[0]
            file_name = varFixedPath[1:] if varDrive else None   
    if option==3:
        #https://stackoverflow.com/questions/48263317/selenium-python-waiting-for-a-download-process-to-complete-using-chrome-web
        sum_before=-1
        sum_after=sum([os.stat(file).st_size for file in os.listdir(activedir)])
        while sum_before != sum_after or any('crdownload' in x for x in os.listdir(activedir)):
            time.sleep(1)
            sum_before = sum_after
            sum_after = sum([os.stat(file).st_size for file in os.listdir(activedir)])
        file_name = get_latest_file(activedir)
    return file_name
            
def set_download_tempdir(activedir):
    D1 = r'\\'+datetime.datetime.now().strftime("tmp_%y%m%d_%H%M%S_%f")
    tempdir=activedir + D1
    if not os.path.exists(tempdir): os.makedirs(tempdir)        
    return tempdir

def otm_download(active_dir,sql,var=None,prefix='',chromePath=None,isdownload=True): #sql : sql script or path+file
    #headless download https://medium.com/@moungpeter/how-to-automate-downloading-files-using-python-selenium-and-headless-chrome-9014f0cdd196
    from selenium.webdriver.common.keys import Keys
    filename='' ;error=''
    download_option=1
    active_dir=active_dir+os.path.sep if os.path.splitdrive(active_dir)[0]==active_dir else active_dir
    tempdir= set_download_tempdir(active_dir) if download_option==1  and isdownload else active_dir
    driver=newChromeBrowser(False,tempdir,chromePath)    
    try:
        #driver.implicitly_wait(3)
        driver.maximize_window()
        driver.get('https://otmgtm-a522235.otm.em2.oraclecloud.com/GC3/glog.webserver.sql.SqlServlet') #"https://otmgtm-test-a522235.otm.em2.oraclecloud.com"
        driver.find_element_by_id("username").send_keys(otm_user)
        driver.find_element_by_id("password").send_keys(otm_password)
        driver.find_element_by_id("signin").click()
        # time.sleep(5)
        # driver.get_screenshot_as_file("capture.png")
        ##driver.find_element_by_id("sql").send_keys(sql)
        if '.sql' in sql and len(sql)<250: 
            if not os.path.isfile(sql): return '','Error: no file:{}'.format(sql)
            sql=otm_query(sql,var)

        import xerox #pip install xerox
        xerox.copy(sql)
        os.system("echo %s| clip" % sql.strip())

#         driver.switch_to.frame(0)
#         # driver.find_element_by_id("userNameId::icon").click() 
#         driver.find_element_by_id("toSpreadsheet").click()
#         driver.find_element_by_name("count@PRF").clear()
#         driver.find_element_by_name("count@PRF").send_keys("99999")
#         driver.find_element_by_id("sql").send_keys(Keys.CONTROL,'v')
#         driver.find_element_by_xpath("//button[@class='enButton']").click()
#         filename=downloads_done(tempdir,download_option,driver)
        if not isdownload:
            driver.switch_to.frame(0)
            #driver.find_element_by_name("count@PRF").clear()
            #driver.find_element_by_name("count@PRF").send_keys("99999")
            driver.find_element_by_id("sql").send_keys(Keys.CONTROL,'v')
            # driver.find_element_by_id("sql").send_keys(sql)
            driver.find_element_by_xpath("//button[@class='enButton']").click()
            time.sleep(180)
            return '','Not error but show result in UI'
        TryTime=1
        while not filename and TryTime<=2:
            if TryTime==2: 
                time.sleep(1)
                print(f'Clear & Retry loading...... \n{tempdir}')
                for filename in os.listdir(tempdir):
                    file_path = os.path.join(tempdir, filename)
                    if os.path.isfile(file_path) or os.path.islink(file_path):
                        os.unlink(file_path)
                driver.refresh()
            driver.switch_to.frame(0)
            if isdownload:
                driver.find_element_by_id("toSpreadsheet").click()
                driver.find_element_by_name("count@PRF").clear()
                driver.find_element_by_name("count@PRF").send_keys("99999")
            driver.find_element_by_id("sql").send_keys(Keys.CONTROL,'v')
            # driver.find_element_by_id("sql").send_keys(sql)
            driver.find_element_by_xpath("//button[@class='enButton']").click() # driver.find_element_by_id("userNameId::icon").click()
            filename=downloads_done(tempdir,download_option,driver)
            TryTime+=1
        #print(os.path.join(active_dir,filename),os.path.abspath(os.path.join(active_dir,filename))) *** abspat not equal to active dir if drive D:
        if not filename: error='\nFail in otm_download of {}, already tried 2 times for 8 minutes each'.format(sql if len(sql)<250 else 'sql query')
        file_name=os.path.splitext(filename)[0] ;file_extension=os.path.splitext(filename)[1]
        if len(filename)<300 and (prefix or not file_extension):
            try:
                filename_2b=prefix[:30-len(file_name)] +('_' if prefix else '')+filename+('' if file_extension else '.xls')
                #print(os.path.split(path_file)[-1],os.path.join(active_dir,filename_))
                shutil.move(os.path.join(active_dir,filename),os.path.join(active_dir,filename_2b)) #+filename[filename.find('.'):]
                filename=filename_2b
                print('Change name to '+filename_2b)
            except Exception as e: print('Cannot change name to'+filename_2b+'::',e) #must not set error=e
    
    except Exception as e:
        error='\nCannot Load OTM in::'+str(var)+str(e)
    
    time.sleep(1)
    driver.close()
    driver.quit()
    print(error)        
    return '' if filename=='' else os.path.join(active_dir,filename),error

def excelBlobCombine(path,filelistname='',storageaccount='scccsupplychainstorage'):
    """Print list of excel blobs in given containter, return combine excel files"""
    #https://pypi.org/project/azure-storage-blob/
    # https://stackoverflow.com/questions/61859634/transform-xlsx-in-blob-storage-to-csv-using-pandas-without-downloading-to-loca
    # https://stackoverflow.com/questions/62370427/read-xlsx-from-azure-blob-storage-to-pandas-dataframe-without-creating-temporary
    from io import BytesIO ,StringIO
    from azure.storage.blob import BlobClient, BlobServiceClient, ContainerClient, __version__ #pip install azure-storage-blob
    #container = ContainerClient.from_connection_string(conn_str=connectionstring, container_name=containername)
    if isinstance(path,str): 
        containername=path
        connectionstring,accountkey=get_blob_connectionstring(containername,storageaccount,r'D:\Siam City Cement Public Company Limited\Logistics_Reports - Documents\script\config.ini')
        container=ContainerClient.from_connection_string(connectionstring,containername)
    else: 
        container=path
        containername=container.get_container_properties().name
        #connectionstring,accountkey=get_blob_connectionstring(containername,storageaccount) --> config.ini may not in drive D:\
    
    if isinstance(filelistname,str): file_list = [file.name for file in container.list_blobs() if filelistname.lower() in file.name.lower()]
    else: file_list = filelistname
    #------------------------------------------------------------
    df = pd.DataFrame() ;allExcelBlobs = [] ;option=1
    #blob_service_client = BlobServiceClient.from_connection_string(connectionstring) 
    for filename in file_list:
        if filelistname.lower() in filename.lower():
            print('blob;',filename)
            extention=os.path.splitext(filename)[1].lower()
            allExcelBlobs.append(filename)
            
            ## blob_client=BlobClient.from_connection_string(conn_str=connectionstring, container_name=containername,blob_name=filename)
            #blob_client = blob_service_client.get_blob_client(container=containername,blob=filename) 
            blob_client = container.get_blob_client(filename)
            ## container_client=blob_service_client.get_container_client(CONTAINERNAME) ;blob_client = container_client.get_blob_client(filename)
            if option==1:
                streamdownloader=blob_client.download_blob()
                if extention in ['.xlsx']:
                    #stream = BytesIO() ;streamdownloader.download_to_stream(stream) ;df_sub = pd.read_excel(stream)
                    #df_sub = pd.read_excel(BytesIO(streamdownloader.content_as_text(encoding=None)))
                    blob=streamdownloader.content_as_bytes()
                    df_sub = pd.read_excel(BytesIO(blob))              
                    if isinstance(df_sub,dict): df_sub=df_sub[0]
                elif extention in ['.csv','.xls']:
                    blob=streamdownloader.content_as_text()
                    df_sub = pd.read_csv(StringIO(blob))     
            
            elif option==2:
                with open("./"+filename,"wb") as my_blob:
                    streamdownloader=blob_client.download_blob()
                    streamdownloader.readinto(my_blob)
                if extention in ['.xlsx']: df_sub=pd.read_excel(filename)
                elif extention in ['.csv','.xls']: df_sub=pd.read_csv(filename)
            
            df = df.append(df_sub,ignore_index=True,sort=False) 
    print('load from blob;{}'.format(allExcelBlobs))
# # local_file_name_out = "csv/prova.csv"
# # container_name_out = "input"

# # blob_client = blob_service_client.get_blob_client(
# #     container=container_name_out, blob=local_file_name_out)
# # blob_client.upload_blob(df.to_csv(path_or_buf = None , encoding='utf-8-sig', index=False))
    return df

# import os ;import Tea as t
# from azure.storage.blob import BlobClient, BlobServiceClient, ContainerClient #pip install azure-storage-blob
# containername='otm-data' ;storageaccount='scccsupplychainstorage' ;connectionstring,accountkey=t.get_blob_connectionstring(containername,storageaccount)
# container=ContainerClient.from_connection_string(connectionstring,containername)
# #print(container.get_container_properties())
# filename='SHIPMENT_REPORT.xlsx'
# blob_service_client = BlobServiceClient.from_connection_string(connectionstring) 
# blob_client = blob_service_client.get_blob_client(container=containername,blob=filename) 

def load_data_(path, filelistname, sheetname=None, skiprows=0,csv_sep=',',is_setdf=False,**read_excel_kwargs): 
    #read_exce :,converters={'names':str,'ages':str} ||| read_csv :,dtype={csv_id:str} or dtype=str If converters are specified, they will be applied INSTEAD of dtype conversion.
    #if sheetname=None, df will be data dict, sheetname=0 --> first sheet
    import os
    import glob
    import collections
    import pandas as pd   
    os.chdir(path)
    df=None
    #cwd = os.getcwd() # Retrieve current working directory (`cwd`)
    #os.listdir()  # List all files and directories in current directory
    filelist = glob.glob('*{}*'.format(filelistname)) if type(filelistname)==str else filelistname
    #is_csvfile = any('.csv' in name.lower() for name in filelist)  
    filelist=[x for x in filelist if x[0].isalnum()]
    if len(filelist)<1: 
        print('No file',os.path.join(path,filelistname))
        return None
#     #SORT FILES: leave only regular files, insert creation date
#     #NOTE: on Windows `ST_CTIME` is a creation date  #  but on Unix it could be something else
#     #NOTE: use `ST_MTIME` to sort by a modification date
# #     dirpath = sys.argv[1] if len(sys.argv) == 2 else r'.'
# #     # get all entries in the directory w/ stats
# #     # entries = (os.path.join(dirpath, fn) for fn in os.listdir(dirpath))
# #     os.chdir(dirpath)
# #     entries = (fn for fn in os.listdir(dirpath))
#     from stat import S_ISREG, ST_CTIME, ST_MTIME, ST_MODE
#     filelist=((os.stat(path), path) for path in filelist) #ctime=datetime.date.fromtimestamp(os.path.getctime(active_dir+r'\\'+filename))
#     filelist=((stat[ST_CTIME], path) for stat, path in filelist if S_ISREG(stat[ST_MODE]))
#     filelist=[path for cdate, path in sorted(filelist)] #os.path.basename(path)
#     if len(filelist)>1: print('filelist:',filelist)
    if sheetname=='': sheetname=None
    pre_sheetname=sheetname
    for x in ['']:
        if x in read_excel_kwargs: read_excel_kwargs.pop(x)
    for idx, filename in enumerate(filelist):
        is_csv=any(x==filename.lower()[-4:] for x in ('.csv','.xls'))
        if is_csv:            
#             if csv_id=='': dfsub=pd.read_csv(filename,skiprows=skiprows,sep=csv_sep)
#             elif csv_id=='str': dfsub=pd.read_csv(filename,skiprows=skiprows,sep=csv_sep,dtype=str) #all columns
#             else : dfsub=pd.read_csv(filename,skiprows=skiprows,sep=csv_sep,**read_excel_kwargs) #specific column name
            csv_sep='\t' if '.xls' in filename.lower() else csv_sep
            dfsub=pd.read_csv(filename,skiprows=skiprows,sep=csv_sep,low_memory=False,**read_excel_kwargs)
#             try: dfsub=pd.read_csv(filename,skiprows=skiprows,sep=csv_sep,low_memory=False,**read_excel_kwargs)
#             except Exception as e: 
#                 print('Cannot load:',filename,e)
#                 return None
            #,dtype={csv_id:str}
            #dfsub=pd.read_table(filename, encoding='utf-16') 
            dfsub.columns = ['' if 'Unnamed' in str(c) else c for c in dfsub]
            #df=df.loc[:, ~df.columns.str.contains('^Unnamed')] #datadict['tee']=df
        else:
            sheet_list=get_all_sheets(path,filename,'')
            if len(sheet_list)==1: sheetname=sheet_list[0] # and len(filelist)==1
            try: dfsub=pd.read_excel(filename,sheet_name=sheetname,skiprows=skiprows,na_values=['NA','nan'],**read_excel_kwargs)
            except Exception as e: 
                print('Cannot load:',filename,e)
                return None
        #keep leading zero when export to excel : df.A = df.A.apply('="{}"'.format)
        #xls = pd.ExcelFile(r"\\".join((path,filename))) #pd.ExcelFile(path+'\\'+filename).sheet_names
        if isinstance(dfsub,pd.DataFrame):
            if dfsub.empty: 
                print('emty data of file',filename)
                is_dataframe=True
                #continue
        if not isinstance(dfsub,pd.DataFrame) and (is_setdf or len(filelist)==1 and isinstance(dfsub,dict) and len(dfsub)==1):
            #*** and not is_csv
            dfsub_=pd.DataFrame()
            for shtname in dfsub.keys(): dfsub_=dfsub_.append(dfsub[shtname],sort=False)
            dfsub=dfsub_
        if idx==0:
            is_dataframe=is_setdf or (isinstance(dfsub,pd.DataFrame) and (len(filelist)==1 or sheetname is not None))
            df=pd.DataFrame() if is_dataframe else collections.OrderedDict()
            #OrderedDict([('a','1'),('b','2')]), ddict=dict(OrderedDict) -->convert order dict to data dict         
        if is_dataframe:
#             dfsub=dfsub.loc[:,~dfsub.columns.astype(str).contains('^Unnamed')]
            dfsub=dfsub.loc[:,~dfsub.columns.str.contains('^Unnamed')] 
            for colname in dfsub.columns: #is_setdf to protect unwanted column (filename)
                if 'unnamed' in str(colname).lower(): del dfsub[colname]     
            if (is_csv or is_setdf) and len(filelist)>0 and not dfsub.empty:
                dfsub.loc[:,'filename']=os.path.splitext(filename)[0]
                #filename.replace('.xls','').replace('.XLS','').replace('.xlsx','').replace('.XLSX','')
            df=df.append(dfsub,sort=False) 
            
        else : df.update({filename.replace('.csv',''): dfsub} if is_csv else dfsub) #df[filename]=dfsub
        
        print('Load '+str(len(dfsub))+(' rows' if is_dataframe else ' sheets')+' of : ' +filename)        
        sheetname=pre_sheetname        
    if len(filelist)>1:
        txt_filelistname=filelistname if type(filelistname)==str else ('[%s]' % ', '.join(map(str, filelistname)))
        print('==>Total ' + str(len(df))+(' rows' if is_dataframe else ' sheets')+' of : ' +txt_filelistname+'\n')
    if isinstance(df,pd.DataFrame): df=filter_columns(df.replace('\'','',regex=True))
    return df
'''After load if data frame
df = load_data(activedir, filename_TruckMaster)
items = list(df.items())
for i in range(len(items)):
    filename = items[i][0]+'.csv'
    items[i][1].to_csv(activedir+'\\'+filename, index=False,encoding = "utf-16") #, sep='\t', encoding='utf-8'
    print('Save file: '+filename)'''

def get_df_from_df_latest_file(active_dir,input_file=None,**read_excel_kwargs): #input_file can be list
    import pandas as pd
    print('#Convert latest file================================================')
    if input_file is None or isinstance(input_file,str): #  or input_file=='' and '.xlsx' not in input_file.lower()):
        df=convert_xls_from_OTM(active_dir,input_file,**read_excel_kwargs)
    #elif isinstance(input_file,str) : df=load_data(active_dir, input_file, None, 0,',',False)
    elif isinstance(input_file,list) and len(input_file)==2: df=load_data(active_dir, input_file[0], input_file[1]) #Active dir, file name, sheet name -> for Silo sensor
    elif isinstance(input_file,pd.DataFrame): df=input_file
    else: 
        print('No data input')
        df=None
    #if sql_table=='ShipmentTracking_Temp': df['Delivery Order No']=df['Delivery Order No'].astype(str) #.apply(str)
    return df

def DateTime_ToString(df,format_str='%Y-%m-%d %H:%M:%S'):
    mycolumns=[x for x in df if 'date' in x.lower()]
    df.reset_index(inplace=True)
    try:
        for col in mycolumns:
            df.loc[df[col].notnull(),col]=pd.to_datetime(df[col]).dt.strftime(format_str)
        print('Convert date column to',format_str)
    except Exception as e: print('Did not convert date column to {}::{}'.format(format_str,e))
    #return df

def ToDateTime(df,format_str='%Y-%m-%d %H:%M:%S'):
    mycolumns=[x for x in df if 'date' in x.lower()]
    for col in mycolumns:
        df.loc[df[col].notnull(),col]=pd.to_datetime(df[col],format=format_str)
    #df.loc[:,mycolumns]=df.loc[:,mycolumns].apply(lambda x:None if pd.isna(x) else pd.to_datetime(x).dt.strftime('%Y-%m-%d %H:%M:%S'))
    return df

def save_data(df, path, filename, sheetname):
    import os
    import datetime
    import pandas as pd
    
    os.chdir(path)
    date_string=datetime.datetime.strftime(datetime.date.today(),'%y%m%d')
    filename=path+'\\'+filename +'_'+ date_string +'.xlsx'
    writer=pd.ExcelWriter(filename,engine='xlsxwriter') #writer = pd.ExcelWriter(path, engine = 'openpyxl')
    df.to_excel(writer, sheet_name=sheetname)
    writer.save()
    writer.close()
    print('\nSave to: '+filename+' --> sheet:'+sheetname)
    
def collectdata_fromfiles(path, filelistname,filename_output):
    import os
    import glob  
    import pandas as pd
    
    os.chdir(path)
    #cwd = os.getcwd() # Retrieve current working directory (`cwd`)
    #os.listdir()  # List all files and directories in current directory
    filelist = glob.glob(filelistname)
    writer=pd.ExcelWriter(path+'\\'+filename_output,engine='xlsxwriter')
    count = 0
    for filename in filelist:
        if filename[0].isalpha():
            dfsub=pd.read_excel(filename,sheetname=0) #read_excel(filename,sheetname = month,skiprows = 12)
            print(str(len(dfsub)) + ' rows : ' +filename)   
            name = filename[:filename.find(' ') if filename.find(' ')!=-1 else filename.find('.')][:30]
            dfsub.to_excel(writer, name)
            count = count+1
    if count>1:
        print('==>Total ' + str(count) + ' files from ' + filelistname + '\n')
        
    writer.save()
    writer.close()
    
def get_all_files(path, fileslist_name ='*.xlsx', filelist_opt='begin_with_char'):
    import os
    import glob
    import pandas as pd
    
    files = []
    os.chdir(path)
    #cwd = os.getcwd() # Retrieve current working directory (`cwd`)
    #os.listdir()  # List all files and directories in current directory
    fileslist = glob.glob(fileslist_name)
    #print(fileslist)
    for filename in fileslist:
        #filename=filename.replace('.xlsx','')
        if filelist_opt=='all' or filename[0].isalnum() :
            files.append(filename)
        else : continue
    print(files)
    return files

def get_all_sheets(path, filename, sheetname=''):
    import pandas as pd
    sheetlist=pd.ExcelFile(path+'\\'+filename).sheet_names
    #sheetlist=load_workbook(path+'\\'+filename).sheetnames
    if sheetname is None: sheetname=''
    return [name for name in sheetlist if sheetname in name]

"""
def get_all_sheets(excel_file, filelist_opt='all'):
    sheets = []
    workbook = load_workbook(excel_file,data_only=True)
    all_worksheets = workbook.sheetnames
    for worksheet_name in all_worksheets:
        sheets.append(worksheet_name)
    return sheets
"""

def get_df_from_sheet(path, filename,sheetname = 0,header_row = 0):
    import pandas as pd   
    xls=pd.ExcelFile(path+'\\'+filename)
    return xls.parse(sheetname, header=header_row, sort=True) #,header = 1, parse_cols = 'B:D'
   
def create_datadict_from_datasheets_in_file(path, filesname_input, sheetname=None, skiprows=0):
    datadict=load_data(path, filesname_input, sheetname, skiprows)
    first_txt=''
    third_txt="EXEC SQL ALTER SESSION SET NLS_DATE_FORMAT = 'YYYYMMDDHH24MISS'"
    remove_list=[]
    for name in datadict.keys():
        first_txt=name
        if first_txt[0].isalpha() :
            df1=datadict[first_txt]
            df1.columns=set_df_multilevel_col(df1,first_txt,third_txt)
        else : remove_list.append(first_txt)
    
    for n in remove_list: del datadict[n]
    print(remove_list)
    return datadict

def write_sheets_to_file_from_datadict(path, filename, data_dict ,truncate_sheet=False,**to_excel_kwargs): # wb : existing file workbook
    #https://stackoverflow.com/questions/49519696/getting-attributeerror-workbook-object-has-no-attribute-add-worksheet-whil?rq=1
    import datetime
    import pandas as pd
    from openpyxl import load_workbook
    short_filename=os.path.splitext(filename)[0]
    filename=short_filename+'.xlsx'
    #short_filename=filename.replace(filename[filename.find('.'):],'') if filename.find('.')>0 else filename
    if isinstance(data_dict,pd.DataFrame) : 
        datadict = {} 
        datadict[short_filename]=data_dict
        data_dict=datadict
    try:
        IsFileExisting=True
        eng='openpyxl' #if os.path.splitext(filename)[1]!='.xls' else 'xlsxwriter'
        pathfile=os.path.abspath(os.path.join(path,filename))
        print(pathfile)
        writer = pd.ExcelWriter(pathfile,engine=eng)
        writer.book = load_workbook(pathfile) #if IsFileExisting :
    except Exception as e: #FileNotFoundError: # file does not exist yet, we will create it
        IsFileExisting=False
        eng='xlsxwriter' #pip install xlsxwriter
        date_string=datetime.datetime.strftime(datetime.date.today(),'%Y-%m-%d')
        filename=short_filename+ '_'+ date_string + '.xlsx' #.replace('\.0', '', regex=True)
        writer = pd.ExcelWriter(path+r'\\'+filename,engine=eng)
        pass
    
    wb=writer.book
    for x in ['engine','header','encoding']:
        if x in to_excel_kwargs: to_excel_kwargs.pop(x)
    for name in data_dict.keys():
        name=os.path.splitext(name)[0]
        startrow=0
        header=True
        #print('startrow' in to_excel_kwargs.keys())
        sheetname=name[:31]
        if sheetname in wb.sheetnames :
            if 'startrow' in to_excel_kwargs.keys():
                startrow=to_excel_kwargs['startrow'] 
                to_excel_kwargs.pop('startrow') 
                if startrow==0: header=data_dict[name].columns
            else: 
                startrow=wb[sheetname].max_row
                header=False
            if truncate_sheet or data_dict[name].index.nlevels>1: 
                #del wb[name] #or wb.remove(wb[name])
                startrow=0
                header=True
                idx=wb.sheetnames.index(sheetname) #idx=workbook.get_sheet_by_name('Sheet2')
                wb.remove(wb.worksheets[idx]) #workbook.remove_sheet(idx)
                wb.create_sheet(sheetname, idx) #wb.create_sheet(index=0,title=name)
                print('Delete sheet :',sheetname)
                
            writer.sheets={ws.title:ws for ws in wb.worksheets} #copy existing sheets        
        if 'index' in to_excel_kwargs:
            isIndex=to_excel_kwargs['index']
            to_excel_kwargs.pop('index')
        else: isIndex=True if data_dict[name].index.nlevels>1 else False #multi index case must set index=True
        data_dict[name].to_excel(writer,sheet_name=sheetname ,encoding="utf-16",header=header,startrow=startrow,index=isIndex,**to_excel_kwargs)
        
        print('Create sheet :'+sheetname+' , '+str(len(data_dict[name]))+' records')
#     # Given a dict of dataframes, for example:
#     # dfs = {'gadgets': df_gadgets, 'widgets': df_widgets}

#     writer = pd.ExcelWriter(filename, engine='xlsxwriter')
#     for sheetname, df in dfs.items():  # loop through `dict` of dataframes
#         df.to_excel(writer, sheet_name=sheetname)  # send df to writer
#         worksheet = writer.sheets[sheetname]  # pull worksheet object
#         for idx, col in enumerate(df):  # loop through all columns
#             series = df[col]
#             max_len = max((
#                 series.astype(str).map(len).max(),  # len of largest item
#                 len(str(series.name))  # len of column name/header
#                 )) + 1  # adding a little extra space
#             worksheet.set_column(idx, idx, max_len)  # set column width
 
    writer.save()              
    if not IsFileExisting :writer.close()
    print('--->Save file: '+filename+'\n')
    return filename

def write_addedsheets_to_file(path, filename, sheetname, sheetlist_out):
    from openpyxl import load_workbook
    wb = load_workbook(path+'\\'+filename)
    writer = pd.ExcelWriter(path+'\\'+filename, engine = 'openpyxl')#(path, engine = 'xlsxwriter')
    writer.book = wb
    sheetlist = wb.sheetnames
    
    data_dict=create_truck_datadict_from_datasheet_in_file(path, filename, sheetname, sheetlist_out)
    #print(data_dict.keys())
    for name in data_dict.keys():
        if name in sheetlist: wb.remove(wb[name]) #or del wb[sheetname] 
        data_dict[name].to_excel(writer, sheet_name = name) #wb.create_sheet(index=0,title=name)
        print('Create sheet :'+name)
        
    writer.save()              
    writer.close()
    print('--->Save file: '+filename+'\n')
    return data_dict

def write_excel_from_csv(activedir,fileinput, input_pathlist):
    print(len(pathlist))
    if len(pathlist)>0:
        for name in pathlist:
            path=activedir+r'\\'+name
            datadict=load_data(path, fileinput)#fileslist=get_all_files(activedir, '*.csv')
            #print(path, '_'+name.replace('\\','_').replace(' ','_')[:30]+'.xlsx')
            write_sheets_to_file_from_datadict(path, '_'+name.replace('\\','_').replace(' ','_')+'_AllCSV'+'.xlsx', datadict)
    else:
            path=activedir
            datadict=load_data(path, fileinput)
            write_sheets_to_file_from_datadict(path, '_test_'.replace(' ','_')+'.xlsx', datadict)
            
def write_csv_from_excel(path, filename, sheets):
    import pandas as pd
    
    xls=pd.ExcelFile(path+'\\'+filename)
    #sheet_to_df_map = {}
    for sheet_name in sheets: #for x in range(0, len(sheets)): 
        try:
            df = xls.parse(sheet_name) #,,header = [0,1],header = 1, parse_cols = 'B:D'
            df.columns = ['' if 'Unnamed' in c else c for c in df]
            df.dropna(axis=0, how='all', inplace=True)
            #sheet_to_df_map[sheet_name] = a #a["Sheet Name"] = [xls.sheet_names[x]] * len(a)
        except KeyError:
            print("Could not find " + sheet_name)
            sys.exit(1)
            
        csvname = ''.join([sheet_name,'.csv'])
        df.to_csv(''.join([path,'\\',csvname]), index=False,encoding = "utf-8") #, sep='\t', encoding='utf-8'
        print('Save file: '+csvname)
        
    print('--->Save file: '+str(len(sheets))+' files\n')
#write_csv_from_excel(activedir, filename, get_all_sheets(activedir, filename,'POWER_UNIT_REMARK'))

def write_csv_from_datadict(path, datadict):
    for name in datadict.keys():
        csvname = ''.join([name,'.csv'])
        datadict[name].to_csv(''.join([path,'\\',csvname]), index=False,encoding="utf-8") #, sep='\t', encoding='utf-8'
        print('Save file: '+csvname)
    print('--->Save file: '+str(len(datadict))+' files\n')

def write_textfile(path, filename, textlist, commandtxt=''):
    txtfile=os.path.abspath(os.path.join(path,filename))#''.join([path,'\\',filename])
    file = open(txtfile,'w',encoding="utf-8")
    txt=''; iscomplete=False
    if isinstance(textlist,str): txt=textlist
    else:
        print('text list, writing ...')
        for t in textlist:
            txt = txt+('\n' if txt else '')+'-dataFileName '+t+' -command '+commandtxt
    try:
        file.write(txt) #txt.encode('utf8')
        file.close() 
        iscomplete=True
        print('save file: '+txtfile)
    except Exception as e: print('CANNOT save {}:: {}'.format(txtfile,e))
    return iscomplete
    
def zipzip(src, dst, filelist=''): #!/usr/bin/env python2.7 #zip is reserve word
    import os
    import zipfile
    import datetime
    
    date_string=datetime.datetime.strftime(datetime.date.today(),'%y%m%d')
    zf = zipfile.ZipFile("%s.zip" % (src+'\\'+dst+'_'+date_string), "w", zipfile.ZIP_DEFLATED)
    abs_src = os.path.abspath(src)
    for dirname, subdirs, files in os.walk(src):
        for filename in files:
            if filename in filelist:
                absname = os.path.abspath(os.path.join(dirname, filename))
                arcname = absname[len(abs_src) + 1:]
                print('zipped file %s' % (arcname)) #os.path.join(dirname, filename),
                zf.write(absname, arcname)
    zf.close()

def set_df_multilevel_col(df1,first_txt,third_txt):
    df1=df1.dropna(axis=0, how='all')
    col_len=len(df1.columns)
    col1=['' for x in range(col_len-1)]
    col1.insert(0,first_txt)
    col3=['' for x in range(col_len-1)]
    col3.insert(0,third_txt)
    return ([col1, df1.columns, col3]) 
    #return 3 records of modified columns 

def get1st_ofgroup(series, prefix='000'):
    print(prefix if len(prefix)>0 else '' +series)
    return series.apply(str).iloc[0] 
def get2nd_ofgroup(series): #rreturn [group['Default Carrier Code'].count(),2]
    return '000'+series.apply(str).iloc[1] if series.count()>=2 else None   
def get3rd_ofgroup(series):
    return '000'+series.apply(str).iloc[2] if series.count()>=3 else None
def get4th_ofgroup(series, prefix='000'):
    return prefix if len(prefix)>0 else ''+series.apply(str).iloc[3] if series.count()>=4 else None
def concat_ofgroup(series):
    return ','.join(map(str, series))
def getorder_ofgroup(series, opt, prefix='000'):
    if opt==2 and series.count()>=2:
        return prefix if len(prefix)>0 else ''+series.apply(str).iloc[1]
    elif opt==3 and series.count()>=3:
        return prefix if len(prefix)>0 else ''+series.apply(str).iloc[2]
    else : return None
    #rreturn [group['Default Carrier Code'].count(),2]
    
#grouped = df.groupby('metacol1')
#transformed = grouped.transform({'col1' : do_something,
#                                'col2' : do_something_else})

# Define the diff function to show the changes in each field
def report_diff(x):
    return x[0] if x[0] == x[1] else '{} ---> {}'.format(*x)

def format_date(date,format_str):
    import datetime
    return datetime.datetime.strftime(date,format_str)
def format_date1(date,format_str):
        import datetime
        import numpy as np
        try:
            return datetime.datetime.strptime(date,format_str)
        except:
            return np.NaN

#=Manage files and folder ===========================================================    
def get_last_x_modified_files(num_files, directory):
    """
    gets a list of files sorted by modified time
    keyword args:
    num_files -- the n number of files you want to print
    directory -- the starting root directory of the search
    """
    import os
    import stat
    import datetime as dt
    import argparse
    modified = []
    accessed = []
    active_dir=os.getcwd()
    active_dir=active_dir+os.path.sep if os.path.splitdrive(active_dir)[0]==active_dir else active_dir
    rootdir=os.path.join(active_dir,directory)
    print(os.getcwd())
    print(rootdir)
    for root, sub_folders, files in os.walk(rootdir):
        for file in files:
            try:
                unix_modified_time = os.stat(os.path.join(root, file))[stat.ST_MTIME]
                unix_accessed_time = os.stat(os.path.join(root, file))[stat.ST_ATIME]
                human_modified_time = dt.datetime.fromtimestamp(unix_modified_time).strftime('%Y-%m-%d %H:%M:%S')
                human_accessed_time = dt.datetime.fromtimestamp(unix_accessed_time).strftime('%Y-%m-%d %H:%M:%S')
                filename = os.path.join('', file) #os.path.join(root, file)
                modified.append((root, filename))
                accessed.append((root, filename, human_accessed_time))
            except:
                pass
    modified.sort(key=lambda a: a[0], reverse=True)
    accessed.sort(key=lambda a: a[0], reverse=True)
    #pprint(accessed[:num_files]) #from pprint import pprint
    if num_files==1:
        for x in range(len(modified)):
            if modified[x][1][0].isalpha():
                #return r"\\".join((modified[x][0],modified[x][1]))
                return modified[x]
    else : return (modified[:num_files])

def delete_ExcelTemp_files():
    python_tempfile = os.environ['USERPROFILE']+r'\\AppData\\Local\\Temp\\gen_py'
    if os.path.exists(python_tempfile):
        print("Found Python Tempfile: {} and delete ...".format(python_tempfile))
        shutil.rmtree(python_tempfile)

def delete_files(dir_list,not_del_list=[],keyword_list=['.xls'],del_list=[],path_date_TupleList=[]):
    if isinstance(dir_list,str): dir_list=[dir_list]
    dir_list_org=dir_list
    dir_list=[folder for folder in dir_list if not any(notdel in folder.lower() for notdel in ['sql','script'])]
    text='ERROR !!! >>> DELETE FOLDER \SQL or \Script' if dir_list_org!=dir_list else ''
    for active_dir in dir_list:
        active_dir=active_dir+os.path.sep if os.path.splitdrive(active_dir)[0]==active_dir else active_dir
        if not os.path.isdir(active_dir): continue
        for filename in os.listdir(active_dir):
            filename=filename.lower()
            if any(r in filename for r in keyword_list) and not any(s.lower() in filename for s in ['tea']+not_del_list):
                delFilename='...\\'+os.path.split(active_dir)[-1]+'\\'+filename
                try:
                    if (not del_list or del_list and any(x.lower() in filename for x in del_list)) \
                            and (not path_date_TupleList \
                                 or path_date_TupleList \
                                    and (not any(tuple_data[0].lower() in active_dir.lower() for tuple_data in path_date_TupleList) \
                                         or any(tuple_data[0].lower() in active_dir.lower() \
                                            and datetime.date.fromtimestamp(os.path.getmtime(fr'{active_dir}\\{filename}')) < tuple_data[1] for tuple_data in path_date_TupleList))):
                        os.remove(os.path.join(active_dir,filename))
                        print('delete:',delFilename)
                except Exception as e: 
                    error=' Cannot delete::{} {}'.format(delFilename,e)
                    text+=error ;print(error)
                    continue
    return text
                
def get_latest_file(activedir,file_input='',extension_name='xls,csv',isalnum=True): #isalnum=True,False,None
    import datetime as dt
    max_file='' ;j = 0 ;file_input=file_input.replace('*','').lower()
    if isinstance(extension_name,str) and ',' in extension_name: extension_name=extension_name.lower().split(',')
    extension_list=extension_name if isinstance(extension_name,list) else [extension_name.lower()]
    if isinstance(activedir,str):
        import stat
        activedir=activedir+os.path.sep if os.path.splitdrive(activedir)[0]==activedir else activedir
        for dirname,subdirs,files in os.walk(activedir):
            for i,fname in enumerate(files):
                isfile=True if isalnum is None else fname[0].isalnum()
                if dirname==activedir and isfile and file_input in fname.lower() and any(word in os.path.splitext(fname)[1].lower() for word in extension_list):
                    full_path=os.path.join(dirname, fname)
                    mtime=os.stat(full_path).st_mtime
                    if j==0: max_mtime=mtime
                    j+=1
                    if mtime >= max_mtime: max_mtime=mtime ;max_file=fname
        if max_file: print('The latest file;'+max_file) #lastmod_file=max_dir, max_file, dt.datetime.fromtimestamp(max_mtime).strftime('%Y-%m-%d %H:%M:%S')
    else: #if isinstance(activedir,ContainerClient)
        from azure.storage.blob import ContainerClient
        container=activedir
        for i,file in enumerate(container.list_blobs()):
            fname=file.name
            isfile=True if isalnum is None else fname[0].isalnum()
            if isfile and file_input in fname.lower() and any(word in os.path.splitext(fname)[1].lower() for word in extension_list):
                mtime=file.last_modified
                if j==0: max_mtime=mtime
                j+=1
                if mtime >= max_mtime: max_mtime=mtime ;max_file=fname
        print('The latest file in blob; {}'.format(max_file if max_file else 'NONE'))
    return max_file

def get_latest_file2(path): #cannot get filename in sub dir.
    import os
    import glob
    #files_list=os.listdir(path)
    files_list=glob.glob(r"\\".join((path,'*')))
    print(len(files_list))
    #return max(files_list, key=os.path.getmtime)

    files = sorted(files_list, key=os.path.getmtime)
    #print ("Oldest:", files[0])
    for x in range(len(files)-1,-1,-1):
        if files[x][0].isalpha():
            return files[x]

def add_history(df,data_list):
    print(data_list) #df.loc[[len(df_his)-1]].to_string(index=False)
    df.loc[len(df)]=data_list
    
def all_files_under(path,wildcard='',isalnum=True,criteria='in',ListType='filename'): #,criteria='under',ListType='pathfile'
    """Iterates through all files that are under the given path."""
    wildcard=wildcard.replace('*','')
    if isinstance(path,str):
        path=path+os.path.sep if os.path.splitdrive(path)[0]==path else path
        for cur_path, dirnames, filenames in os.walk(path):
            if criteria=='in' and cur_path!=path: continue
            for filename in filenames:
                if wildcard.lower() in filename.lower() and (True if isalnum is None else filename[0].isalnum()):
                    yield filename if ListType=='filename' else os.path.join(cur_path,filename)
    else:
        container=path ;containername=container.get_container_properties().name
        for file in container.list_blobs():
            filename=file.name
            if wildcard.lower() in filename.lower() and (True if isalnum is None else filename[0].isalnum()):
                yield filename if ListType=='filename' else os.path.join(containername,filename)
# print(max(all_files_under(path,'in'), key=os.path.getmtime))
#[x for x in all_files_under(path,'daily',True,'in','filename')]

def rename_file(path,origname,tobename,issuffix=False):
    #if issuffix: TimeString=datetime.datetime.strftime(datetime.datetime.now(),'_%m-%d_%H%M')
    TimeString=datetime.datetime.strftime(datetime.datetime.now(),'_%H%M') if issuffix else ''
    origname=get_latest_file(path,origname)
    newname=origname ;error=''
    if len(origname)<300 and origname:
        extension=os.path.splitext(origname)[1]
        target=os.path.splitext(tobename)[0]
        source_digits=30-len(target)-len(TimeString)
        source=os.path.splitext(origname.lower())[0][:-3][:source_digits] #diff in lower, upper may error in invalid argument
        tobename='{}{}_{}{}'.format(source.lower(),TimeString,target,extension) 
        try:
            if isinstance(path,str):
                path=os.path.abspath(path)
                shutil.move(os.path.join(path,origname),os.path.join(path,tobename)) #+filename[filename.find('.'):]
            else:
                container=path
                source_blob = container.get_blob_client(origname)
                copied_blob = container.get_blob_client(tobename)
                
                copy = copied_blob.start_copy_from_url(source_blob.url) #name case sensitive
                props = copied_blob.get_blob_properties()
                print('Blob copied:{} <-- {}'.format(props.copy.status,source_blob.url))
                if props.copy.status=='success': source_blob.delete_blob()
            newname=tobename
        except Exception as e: error=str(e) #must not set error=e        
    else: error='No file exist in {}'.format(path)
    print('Change name from {} to {} {}'.format(origname if len(origname)<300 else '',tobename, ':: Fail, '+error if error else 'success'))
    return newname 

def get_data_fleetcap_(active_dir,output_name='FleetCap'):   
    import gspread    
    import configparser
    from oauth2client.service_account import ServiceAccountCredentials
    from datetime import timedelta, date
      
    ######################## Set Date Time and Template File Name ################################
    D1 = date.today() - timedelta(days=0)
    M1 = D1.strftime("%Y_%m%b")
    D1 = D1.strftime("%d-%m-%Y")

    fday_curmonth = date.today().replace(day=1)
    lday_premonth = fday_curmonth - timedelta(days=1)

    M0 = lday_premonth.strftime("%Y_%m%b")

    pm_file_name = active_dir+r'\\'+output_name+'_Template_'+M0+'.xlsx'
    file_name = active_dir+r'\\'+output_name+'_Template_'+M1+'.xlsx'
    print(file_name)

    ######################## Set Date Time and Template File Name ################################
    if os.path.exists(file_name):
        print("Found Template File") 
    else:
        print("Not Found Template File") 
        df_pm = pd.read_excel(pm_file_name,sheet_name = output_name)
        cols = ['Shipping Type',	'Transporter Code', 'Transporter Name', 'Fleet commit', 'FleetCap']
        df_pm = df_pm[cols]
        df_pm = df_pm.astype(str)
        # write and update
        excel_template_writer = pd.ExcelWriter(file_name)
        df_pm.to_excel(excel_template_writer, output_name, index=False)  
        excel_template_writer.save()
        excel_template_writer.close()
        print('Create New Template File :', file_name)
        
    ########## Read Transporter Code and Shipping Type from Template #########################
        
    df = pd.read_excel(file_name,sheet_name = output_name)
    cols = ['Shipping Type','Transporter Code']
    df = df[cols]
    df = df.astype(str)
    template_shipping_type = df['Shipping Type'].tolist() 
        
    for index, row in df.iterrows():
        status = False
        trancode = row['Transporter Code']
        ########################## Get Data From Google Sheet ####################################
        scope = ['https://spreadsheets.google.com/feeds',
                 'https://www.googleapis.com/auth/drive']
        # Your json file here
        credentials = ServiceAccountCredentials.from_json_keyfile_name('zinc-proton-268015-d2cf843be317.json', scope)
        gc = gspread.authorize(credentials)
        try:
            wks = gc.open(trancode).sheet1
            data = wks.get_all_values()
        except:
            fleetcap.append(0)
            continue

        ########################### Get Fleet Cap Data ##########################################
        headers = data.pop(0)
        df_gd = pd.DataFrame(data, columns=headers)
        df_gd = df_gd[0:]
        
        for i in range(0, len(df_gd)):
                       
            ##print('TemplateShippingType :',template_shipping_type[index])
            new_shipping_type = df_gd['ประเภทรถ'].astype(str)
            ##print('NewShippingType :',new_shipping_type[i])
            commit = df_gd['จำนวนงานที่รับได้'].astype(str).tolist()
            
            ##print('index1',index)
            ##print('index2',i)
            fleet_commit=int(commit[i])
            print(trancode,'-> FleetCommit :',fleet_commit)
            
            if index == 0 and template_shipping_type[index] == new_shipping_type[i]:
                fleetcap = [fleet_commit]
                status = True
            elif template_shipping_type[index] == new_shipping_type[i] :
                fleetcap.append(fleet_commit)
                status = True
                #print(fleetcap)
        if status == False :
            fleetcap.append(0)
    #print(fleetcap)
    ############################# Get Exiting Excel Template #########################################
    # get data to be appended
    df_append = pd.DataFrame({D1: fleetcap})

    # define what sheets to update
    to_update = {"FleetCap": df_append}

    # load existing data
    excel_reader = pd.ExcelFile(file_name)

    # write and update
    excel_writer = pd.ExcelWriter(file_name)

    for sheet in excel_reader.sheet_names:
        sheet_df = excel_reader.parse(sheet)
        append_df = to_update.get(sheet)

        if append_df is not None:

            if D1 in sheet_df.columns:
                sheet_df.update(append_df)
            else:
                sheet_df = pd.concat([sheet_df, df_append], axis=1)
                
            sheet_df.loc[:,'FleetCap']=fleetcap
            sheet_df.loc[:,'Transporter Code']=sheet_df['Transporter Code'].apply(str) 

        sheet_df.to_excel(excel_writer, sheet, index=False)
        
    excel_writer.save()
    excel_writer.close()
    return sheet_df

def get_data_fleetcap(active_dir,output_name='FleetCap'):
    import gspread    
    from oauth2client.service_account import ServiceAccountCredentials
    from datetime import timedelta, date
      
    ######################## Set Date Time and Template File Name ################################
    D1 = date.today() - timedelta(days=0)
    M1 = D1.strftime("%Y_%m%b")
    D1 = D1.strftime("%d-%m-%Y")

    fday_curmonth = date.today().replace(day=1)
    lday_premonth = fday_curmonth - timedelta(days=1)

    M0 = lday_premonth.strftime("%Y_%m%b")

    pm_file_name = active_dir+r'\\'+output_name+'_Template_'+M0+'.xlsx'
    file_name = active_dir+r'\\'+output_name+'_Template_'+M1+'.xlsx'
    #print(file_name)

    ######################## Set Date Time and Template File Name ################################
    if os.path.exists(file_name):
        print("Found Template File") 
    else:
        print("Not Found Template File") 
        df_pm = pd.read_excel(pm_file_name,sheet_name = output_name)
        cols = ['Shipping Type','Sub Shipping Type','Transporter Code', 'Transporter Name', 'Fleet commit', 'FleetCap']
        df_pm = df_pm[cols]
        df_pm = df_pm.astype(str)
        # write and update
        excel_template_writer = pd.ExcelWriter(file_name)
        df_pm.to_excel(excel_template_writer, output_name, index=False)  
        excel_template_writer.save()
        excel_template_writer.close()
        print('Create New Template File :', file_name)
        
    ########## Read Transporter Code and Shipping Type from Template #########################
        
    df = pd.read_excel(file_name,sheet_name = output_name)
    cols = ['Shipping Type','Sub Shipping Type','Transporter Code']
    df = df[cols]
    df = df.astype(str)
    template_shipping_type = df['Shipping Type'].tolist() 
    template_subshipping_type = df['Sub Shipping Type'].tolist()
    for index, row in df.iterrows():
        status = False
        trancode = row['Transporter Code']
        ########################## Get Data From Google Sheet ####################################
        scope = ['https://spreadsheets.google.com/feeds',
                 'https://www.googleapis.com/auth/drive']
        # Your json file here
        credentials = ServiceAccountCredentials.from_json_keyfile_name('zinc-proton-268015-d2cf843be317.json', scope)
        gc = gspread.authorize(credentials)
        try:
            wks = gc.open(trancode).sheet1
            data = wks.get_all_values()
        except:
            fleetcap.append(0)
            continue

        ########################### Get Fleet Cap Data ##########################################
        headers = data.pop(0)
        df_gd = pd.DataFrame(data, columns=headers)
        df_gd = df_gd[0:]
        
        for i in range(0, len(df_gd)):
                       
            ##print('TemplateShippingType :',template_shipping_type[index])
            new_shipping_type = df_gd['ประเภทรถ'].astype(str)
            new_subshipping_type = df_gd['ประเภทรถย่อย'].astype(str)
            commit = df_gd['จำนวนงานที่รับได้'].astype(str).tolist()
            
            ##print('index1',index)
            ##print('index2',i)
            fleet_commit=int(commit[i])
            print(trancode,':',new_subshipping_type[i],'-> FleetCommit :',fleet_commit)
            
            if template_subshipping_type[index] != 'nan':
                if index == 0 and template_shipping_type[index] == new_shipping_type[i]:
                    fleetcap = [fleet_commit]
                    status = True
                elif template_shipping_type[index] == new_shipping_type[i] and template_subshipping_type[index] == new_subshipping_type[i]:
                    fleetcap.append(fleet_commit)
                    status = True
                    #print(fleetcap)                
            else:
                if index == 0 and template_shipping_type[index] == new_shipping_type[i] :
                    fleetcap = [fleet_commit]
                    status = True
                elif template_shipping_type[index] == new_shipping_type[i]:
                    fleetcap.append(fleet_commit)
                    status = True
                    #print(fleetcap)
                    
        if status == False :
            fleetcap.append(0)
    #print(fleetcap)
    ############################# Get Exiting Excel Template #########################################
    # get data to be appended
    df_append = pd.DataFrame({D1: fleetcap})

    # define what sheets to update
    to_update = {"FleetCap": df_append}

    # load existing data
    excel_reader = pd.ExcelFile(file_name)

    # write and update
    excel_writer = pd.ExcelWriter(file_name)

    for sheet in excel_reader.sheet_names:
        sheet_df = excel_reader.parse(sheet)
        append_df = to_update.get(sheet)

        if append_df is not None:

            if D1 in sheet_df.columns:
                sheet_df.update(append_df)
            else:
                sheet_df = pd.concat([sheet_df, df_append], axis=1)
                
            sheet_df.loc[:,'FleetCap']=fleetcap
            sheet_df.loc[:,'Transporter Code']=sheet_df['Transporter Code'].apply(str) 

        sheet_df.to_excel(excel_writer, sheet, index=False)
        
    excel_writer.save()
    excel_writer.close()
    return sheet_df

def get_data_fleetcap_history(active_dir,output_name='FleetCap'):
    import gspread    
    from oauth2client.service_account import ServiceAccountCredentials
    from datetime import timedelta, date
      
    ######################## Set Date Time and Template File Name ################################
    D1 = date.today() - timedelta(days=0)
    M1 = D1.strftime("%Y_%m%b")
    D1 = D1.strftime("%d-%m-%Y")

    file_name = active_dir+r'\\'+output_name+'_Template_'+M1+'.xlsx'
    print(file_name)

    ######################## Set Date Time and Template File Name ################################
    if os.path.exists(file_name):
        print("Found Template File")
        
        ########## Read Transporter Code and Shipping Type from Template #########################
        df = pd.read_excel(file_name,sheet_name = output_name)
        cols = ['Shipping Type','Transporter Code']
        df = df[cols]
        df = df.astype(str)
        #template_shipping_type = df['Shipping Type'].tolist()
        
        for index, row in df.iterrows():
            status = False
            trancode = row['Transporter Code']
            ########################## Get Data From Google Sheet ####################################
            scope = ['https://spreadsheets.google.com/feeds',
                     'https://www.googleapis.com/auth/drive']
            # Your json file here
            credentials = ServiceAccountCredentials.from_json_keyfile_name('zinc-proton-268015-d2cf843be317.json', scope)
            gc = gspread.authorize(credentials)
            try:
                wks = gc.open(trancode).sheet1
                data = wks.get_all_values()
            except:
                #fleetcap.append(0)
                break

            ########################### Get Fleet Cap Data ##########################################
            headers = data.pop(0)
            
            df = pd.DataFrame(data, columns=headers)
            #print(df.head())
            df.insert(6, 'TimeStamp', pd.datetime.now().replace(microsecond=0))
            ############################# Get Exiting Excel Template ##################################
            # load existing data
            file_name = active_dir+'\\FleetCap_History_Template.xls'
            output_name='FleetCap'

            if os.path.exists(file_name):
                print("Found Template File") 
                df2 = pd.read_excel(file_name,sheet_name = output_name)
                df2 = pd.DataFrame(df2)
                #print(df2)
                sheet_df = df.append(df2)
            else:
                print("Not Found Template File")     
                sheet_df = df
             
            #print(sheet_df)
            writer = pd.ExcelWriter(file_name)
            sheet_df.to_excel(writer, output_name, index=False) 
            writer.save()
            writer.close()   
    else:
        print('Not Found Template File > Please Check File : FleetCap_Template'+M1+'.xlsx')
        
    return sheet_df

# ================================================================================================================================
def set_screen(x, y):
    import win32api
    import win32con
    import pywintypes

    devmode = pywintypes.DEVMODEType()

    devmode.PelsWidth = x
    devmode.PelsHeight = y

    devmode.Fields = win32con.DM_PELSWIDTH | win32con.DM_PELSHEIGHT

    win32api.ChangeDisplaySettings(devmode, 0)


def click(x, y):
    from ctypes import windll
    #move first
    windll.user32.SetCursorPos(x, y)

    #then click
    windll.user32.mouse_event(2, 0, 0, 0, 0)
    windll.user32.mouse_event(4, 0, 0, 0, 0)


def start_vpn(cmd):
    from subprocess import Popen
    from ctypes import windll
    from tkinter import Tk

    user32 = windll.user32

    root = Tk()
    root.update()
    root.wm_state('iconic')

    initial_screen_width = int(root.winfo_screenwidth())
    initial_screen_hight = int(root.winfo_screenheight())
    print('initial_screen_width', initial_screen_width)
    print('initial_screen_hight', initial_screen_hight)

    if initial_screen_width == 1024:
        click(995, 560)  # Click Pop-Up Alert
    else:
        click(1335, 560)  # Click Pop-Up Alert

    time.sleep(1)
    #Set Today Date and Format
    Popen('C:\Program Files\Palo Alto Networks\GlobalProtect\PanGPA.exe')

    screen_width = int(root.winfo_screenwidth())-50
    screen_height = int(root.winfo_screenheight())-80

    print('screen_width_point', screen_width)
    print('screen_height_point', screen_height)

    time.sleep(1)
    click(screen_width, screen_height)

    time.sleep(3)
    if initial_screen_width == 1024:
        click(995, 560)  # Click Pop-Up Alert
    else:
        click(1335, 560)  # Click Pop-Up Alert

    root.withdraw()

    print("Start VPN Already.")
    isVpnConnected = True

    set_screen(1366, 768)
    return isVpnConnected

def ping_test_start_vpn(hostname = '10.254.1.181',username=username):
    start_count = 0 ;isVpnConnected=False
    while True:
        time.sleep(1)
        response = os.system('ping ' + hostname)
        if response == 0:
            print(hostname, 'is up')
            isVpnConnected=True
            break
        elif username != 'supplychainvm':
            print(f'********* need to turn on VPN manually befor run DAS **********  (username;{username} is not supplychainvm)')
            isVpnConnected=False
            break
        else:
            print(hostname, 'is down')
            isVpnConnected = start_vpn('start')
            start_count = start_count + 1
            #if start_count > 30:
            #    break
            print(f'is Vpn Connected: {isVpnConnected}')
    return isVpnConnected
# =================================================================================================================
